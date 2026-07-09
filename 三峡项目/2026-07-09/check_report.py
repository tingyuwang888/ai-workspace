#!/usr/bin/env python3
"""
tiance-report-checker v1.0.0
天策策略测试报告自动化质量检查工具

对测试报告 Excel 进行多维度质量检查，输出带标注的新 Excel 文件。
支持两种使用方式：
  1. 独立运行：python3 check_report.py <报告.xlsx> [-o 输出.xlsx]
  2. 作为模块导入：from check_report import check_report

检查维度：
  1. 判定一致性 — 通过/未通过状态与实际结果是否矛盾
  2. 数据完整性 — 空值、编号连续/重复、版本号、时间格式
  3. 语义重复   — 预期结果完全相同、名称高度相似的用例
  4. 实际结果质量 — 函数输出匹配、策略异常检测
"""

import sys
import os
import re
import json
import argparse
from collections import Counter, defaultdict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 需要 openpyxl 库。请运行: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

VERSION = "1.0.0"

# ============================================================
# 列格式自动检测
# ============================================================

_HEADERS_15 = [
    '用例编号', '模块', '用例名称', '业务描述', '测试场景',
    '用例类型', '前置条件', '测试步骤', '预期结果', '实际结果',
    '测试状态', '测试时间', '测试流水号UUID', '测试数据', '三方数据'
]

_RULE_CODE_RE = re.compile(
    r'(ARGP\d+|EBGS\d+|EBGP\d+|EBNP\d+|EBNS\d+|ENDX\d+|ARMP\d+|'
    r'ELLP\d+|ELLS\d+|MAIN\d+|RED_ALERT_\w+|RULE_\w+)'
)

_FUNC_CODE_RE = re.compile(r'(S\d{6}|FUNC_\d+)')

_RISK_LEVELS = ['极高风险', '高风险', '中风险', '低风险', '红色预警', '黄色预警', '蓝色预警', '绿色预警']

_ERROR_KEYWORDS = ['报异常', '异常中断', '执行失败', '策略报异常', '系统异常', '运行异常', '超时']


def detect_format(ws):
    """自动检测报告列格式，返回 (format_name, column_map)"""
    header_cells = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h).strip() if h else '' for h in header_cells]

    # 精确匹配15列
    if headers[:15] == _HEADERS_15:
        return '15col', _build_col_map_15()

    # 模糊匹配：通过关键列名定位
    col_map = {}
    for i, h in enumerate(headers, 1):
        if '用例编号' in h:
            col_map['tc_id'] = i
        elif h == '模块' or h == '策略名称/规则集':
            col_map['module'] = i
        elif '用例名称' in h or ('规则名称' in h and 'tc_name' not in col_map):
            col_map['tc_name'] = i
        elif '业务描述' in h or '规则描述' in h:
            col_map['biz_desc'] = i
        elif '测试场景' in h:
            col_map['scenario'] = i
        elif '用例类型' in h or '案例类型' in h:
            col_map['case_type'] = i
        elif '前置条件' in h or '前提条件' in h:
            col_map['precondition'] = i
        elif '测试步骤' in h or '步骤描述' in h:
            col_map['test_step'] = i
        elif h == '预期结果':
            col_map['expected'] = i
        elif '实际结果' in h or '实际执行结果' in h:
            col_map['actual'] = i
        elif '测试状态' in h or '执行状态' in h:
            col_map['status'] = i
        elif '测试时间' in h or '执行时间' in h:
            col_map['test_time'] = i
        elif 'UUID' in h.upper() or 'Token' in h or '流水号' in h:
            col_map['uuid'] = i
        elif h == '测试数据':
            col_map['test_data'] = i
        elif '三方数据' in h:
            col_map['third_party'] = i

    required = ['tc_id', 'expected', 'actual', 'status']
    if all(k in col_map for k in required):
        for key in ['module', 'tc_name', 'biz_desc', 'scenario', 'case_type',
                     'precondition', 'test_step', 'test_time', 'uuid', 'test_data', 'third_party']:
            col_map.setdefault(key, None)
        return 'auto', col_map

    return 'unknown', {}


def _build_col_map_15():
    return {
        'tc_id': 1, 'module': 2, 'tc_name': 3, 'biz_desc': 4, 'scenario': 5,
        'case_type': 6, 'precondition': 7, 'test_step': 8, 'expected': 9,
        'actual': 10, 'status': 11, 'test_time': 12, 'uuid': 13,
        'test_data': 14, 'third_party': 15,
    }


# ============================================================
# 数据读取
# ============================================================

def read_rows(ws, col_map):
    """读取所有数据行，返回 list of dict"""
    rows = []
    for r in range(2, ws.max_row + 1):
        def val(key):
            col = col_map.get(key)
            if col is None:
                return None
            return ws.cell(row=r, column=col).value
        row = {
            '_row': r,
            'tc_id': str(val('tc_id') or '').strip(),
            'module': str(val('module') or '').strip(),
            'tc_name': str(val('tc_name') or '').strip(),
            'biz_desc': str(val('biz_desc') or '').strip(),
            'scenario': str(val('scenario') or '').strip(),
            'case_type': str(val('case_type') or '').strip(),
            'precondition': str(val('precondition') or '').strip(),
            'test_step': str(val('test_step') or '').strip(),
            'expected': str(val('expected') or '').strip(),
            'actual': str(val('actual') or '').strip(),
            'status': str(val('status') or '').strip(),
            'test_time': str(val('test_time') or '').strip(),
            'uuid': str(val('uuid') or '').strip(),
            'test_data': str(val('test_data') or '').strip(),
            'third_party': str(val('third_party') or '').strip(),
        }
        if row['tc_id'] or row['tc_name'] or row['expected']:
            rows.append(row)
    return rows


# ============================================================
# 维度一：判定一致性
# ============================================================

def check_verdict_consistency(rows):
    issues = []
    for row in rows:
        tc = row['tc_id'] or f"行{row['_row']}"
        status = row['status']
        expected = row['expected']
        actual = row['actual']

        is_pass = status == '通过' or status.startswith('通过')

        # 1. 目标规则预期命中但实际未命中，却标"通过"
        if is_pass:
            target_rules = _extract_target_rules(row)
            for rule in target_rules:
                if f'未命中{rule}' in actual:
                    issues.append({
                        'row': row['_row'], 'tc_id': tc, 'level': '严重',
                        'category': '判定一致性',
                        'issue': f'目标规则{rule}预期命中但实际未命中，状态却为"通过"',
                    })
                    break

        # 2. 策略报异常却标"通过"
        if is_pass:
            for kw in _ERROR_KEYWORDS:
                if kw in actual:
                    issues.append({
                        'row': row['_row'], 'tc_id': tc, 'level': '严重',
                        'category': '判定一致性',
                        'issue': f'实际结果含"{kw}"但状态标为"通过"',
                    })
                    break

        # 3. 通过/未通过但实际结果为空
        if is_pass and not actual:
            issues.append({
                'row': row['_row'], 'tc_id': tc, 'level': '警告',
                'category': '判定一致性',
                'issue': '状态为"通过"但实际结果为空',
            })
        if '未通过' in status and not actual:
            issues.append({
                'row': row['_row'], 'tc_id': tc, 'level': '警告',
                'category': '判定一致性',
                'issue': '状态为"未通过"但实际结果为空',
            })

    return issues


def _extract_target_rules(row):
    """从用例名称和预期结果中提取目标规则编号"""
    rules = set()
    name = row['tc_name']
    expected = row['expected']

    m = _RULE_CODE_RE.search(name)
    if m:
        rules.add(m.group(1))

    hit_rules = _RULE_CODE_RE.findall(expected)
    if '命中' in expected and hit_rules:
        rules.add(hit_rules[0])

    return rules


# ============================================================
# 维度二：数据完整性
# ============================================================

def check_data_integrity(rows, ws, filename=''):
    issues = []

    # 1. 关键列空值
    critical_cols = ['tc_id', 'tc_name', 'expected', 'status']
    for row in rows:
        tc = row['tc_id'] or f"行{row['_row']}"
        for col in critical_cols:
            val = row.get(col, '')
            if not val:
                issues.append({
                    'row': row['_row'], 'tc_id': tc, 'level': '警告',
                    'category': '数据完整性',
                    'issue': f'关键列"{col}"为空',
                })

    # 2. TC编号连续性
    tc_nums = []
    for row in rows:
        m = re.match(r'TC_(\d+)', row['tc_id'])
        if m:
            tc_nums.append(int(m.group(1)))

    if tc_nums:
        tc_set = set(tc_nums)
        min_n, max_n = min(tc_nums), max(tc_nums)
        expected_set = set(range(min_n, max_n + 1))
        gaps = expected_set - tc_set
        if gaps and len(gaps) < len(tc_nums) * 0.3:
            gap_strs = [f'TC_{g:03d}' for g in sorted(gaps)[:10]]
            suffix = f'...等{len(gaps)}个' if len(gaps) > 10 else ''
            issues.append({
                'row': 0, 'tc_id': '(全局)', 'level': '提示',
                'category': '数据完整性',
                'issue': f'编号不连续，缺失{len(gaps)}个: {", ".join(gap_strs)}{suffix}',
            })

    # 3. TC编号重复
    base_ids = []
    for row in rows:
        m = re.match(r'(TC_\d+)', row['tc_id'])
        if m:
            base_ids.append((m.group(1), row))

    dup_counter = Counter(b for b, _ in base_ids)
    for tc_id, count in dup_counter.items():
        if count > 1:
            dup_rows = [r['_row'] for b, r in base_ids if b == tc_id]
            issues.append({
                'row': dup_rows[0], 'tc_id': tc_id, 'level': '严重',
                'category': '数据完整性',
                'issue': f'编号{tc_id}重复出现{count}次（行{dup_rows}）',
            })

    # 4. policyVersion一致性
    filename_ver = _extract_version_from_filename(filename)
    if filename_ver:
        ver_count = 0
        for row in rows:
            full_text = row['precondition'] + ' ' + row['test_step']
            versions = re.findall(r'policyVersion[=:](\d+)', full_text)
            for v in versions:
                if v != filename_ver:
                    ver_count += 1
                    break
        if ver_count:
            issues.append({
                'row': 0, 'tc_id': '(全局)', 'level': '警告',
                'category': '数据完整性',
                'issue': f'文件名标注V{filename_ver}，但{ver_count}条用例引用policyVersion≠{filename_ver}',
            })

    # 5. 测试时间格式不一致
    time_formats = Counter()
    for row in rows:
        t = row['test_time']
        if not t:
            continue
        if re.match(r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}$', t):
            time_formats['YYYY-MM-DD HH:MM'] += 1
        elif re.match(r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$', t):
            time_formats['YYYY-MM-DD HH:MM:SS'] += 1
        elif re.match(r'^\d{4}-\d{2}-\d{2}$', t):
            time_formats['YYYY-MM-DD'] += 1
        else:
            time_formats['其他'] += 1

    if len(time_formats) > 1:
        fmt_str = ', '.join(f'{k}({v}条)' for k, v in time_formats.most_common())
        issues.append({
            'row': 0, 'tc_id': '(全局)', 'level': '提示',
            'category': '数据完整性',
            'issue': f'测试时间格式不统一: {fmt_str}',
        })

    # 6. 用例编号含批注
    for row in rows:
        tc = row['tc_id']
        if '（' in tc or '(' in tc:
            m = re.search(r'[（(](.+?)[）)]', tc)
            comment = m.group(1) if m else ''
            issues.append({
                'row': row['_row'], 'tc_id': tc, 'level': '提示',
                'category': '数据完整性',
                'issue': f'用例编号含人工批注: "{comment[:60]}"',
            })

    # 7. 三方数据列为空（仅对规则类命中用例）
    for row in rows:
        tc = row['tc_id'] or f"行{row['_row']}"
        if row['module'] == '规则' and not row['third_party'] and row['third_party'] != 'None':
            if row['expected'] and '命中' in row['expected']:
                issues.append({
                    'row': row['_row'], 'tc_id': tc, 'level': '提示',
                    'category': '数据完整性',
                    'issue': '规则类用例三方数据列为空',
                })

    return issues


def _extract_version_from_filename(filename):
    if not filename:
        return None
    m = re.search(r'[Vv](\d+)', filename)
    return m.group(1) if m else None


# ============================================================
# 维度三：语义重复检测
# ============================================================

def check_semantic_duplicates(rows):
    issues = []

    # 1. 同模块内预期结果完全相同
    by_module = defaultdict(list)
    for row in rows:
        if row['module'] and row['expected']:
            by_module[row['module']].append(row)

    for module, module_rows in by_module.items():
        exp_groups = defaultdict(list)
        for row in module_rows:
            exp = row['expected'].strip()
            if exp:
                exp_groups[exp].append(row)

        for exp, group in exp_groups.items():
            if len(group) > 1:
                tc_ids = [r['tc_id'] or f"行{r['_row']}" for r in group]
                names = [r['tc_name'][:40] for r in group]
                is_boundary_dup = any('边界' in n for n in names) and any('规则' in n for n in names)
                issues.append({
                    'row': group[0]['_row'], 'tc_id': tc_ids[0],
                    'level': '警告' if is_boundary_dup else '提示',
                    'category': '语义重复',
                    'issue': f'模块"{module}"内{len(group)}条预期结果完全相同: {", ".join(tc_ids[:5])}',
                })

    # 2. 名称高度相似的用例对（>85%）
    seen_pairs = set()
    n = len(rows)
    for i in range(n):
        for j in range(i + 1, min(i + 50, n)):
            name_i = rows[i]['tc_name']
            name_j = rows[j]['tc_name']
            if not name_i or not name_j or len(name_i) < 10 or len(name_j) < 10:
                continue
            ratio = _sequence_similarity(name_i, name_j)
            if 0.85 < ratio < 1.0:
                pair_key = (min(i, j), max(i, j))
                if pair_key not in seen_pairs:
                    seen_pairs.add(pair_key)
                    tc_i = rows[i]['tc_id'] or f"行{rows[i]['_row']}"
                    tc_j = rows[j]['tc_id'] or f"行{rows[j]['_row']}"
                    issues.append({
                        'row': rows[i]['_row'], 'tc_id': tc_i,
                        'level': '提示',
                        'category': '语义重复',
                        'issue': f'与{tc_j}名称相似度{ratio:.0%}（可能跨子策略对应规则）',
                    })

    return issues


def _sequence_similarity(a, b):
    if a == b:
        return 1.0
    len_a, len_b = len(a), len(b)
    if len_a == 0 or len_b == 0:
        return 0.0
    matches = sum(1 for ca, cb in zip(a, b) if ca == cb)
    return matches / max(len_a, len_b)


# ============================================================
# 维度四：实际结果质量
# ============================================================

def check_result_quality(rows):
    issues = []
    for row in rows:
        tc = row['tc_id'] or f"行{row['_row']}"
        expected = row['expected']
        actual = row['actual']
        module = row['module']
        name = row['tc_name']

        if not actual or not expected:
            continue

        # 1. 函数类用例：实际结果中的函数与预期目标不匹配
        if module == '函数' or '函数' in name or _FUNC_CODE_RE.search(name):
            expected_funcs = _FUNC_CODE_RE.findall(expected)
            actual_funcs = _FUNC_CODE_RE.findall(actual)
            if expected_funcs and actual_funcs:
                expected_set = set(expected_funcs)
                actual_set = set(actual_funcs)
                if not expected_set & actual_set:
                    issues.append({
                        'row': row['_row'], 'tc_id': tc, 'level': '严重',
                        'category': '实际结果质量',
                        'issue': f'函数用例：预期目标函数{sorted(expected_set)}但实际结果中记录的函数为{sorted(actual_set)}（可能记录了错误的函数输出）',
                    })

        # 2. 规则类用例：目标规则未命中但有伴随命中
        if module == '规则':
            name_rules = _RULE_CODE_RE.findall(name)
            actual_hit_rules = _RULE_CODE_RE.findall(actual)
            if name_rules and '命中' in expected:
                target = name_rules[0]
                if f'未命中{target}' in actual:
                    other_hits = [r for r in actual_hit_rules if r != target]
                    if other_hits:
                        issues.append({
                            'row': row['_row'], 'tc_id': tc, 'level': '警告',
                            'category': '实际结果质量',
                            'issue': f'目标规则{target}未命中，伴随命中{other_hits[:3]}（可能mock数据问题）',
                        })

        # 3. 风险等级不一致（仅在目标规则未命中时报告）
        exp_levels = [lv for lv in _RISK_LEVELS if lv in expected]
        act_levels = [lv for lv in _RISK_LEVELS if lv in actual]
        if exp_levels and act_levels:
            exp_final = exp_levels[-1]
            act_final = act_levels[-1]
            if exp_final != act_final:
                target_rules = _extract_target_rules(row)
                target_hit = any(f'命中{r}' in actual for r in target_rules)
                if not target_hit and target_rules:
                    issues.append({
                        'row': row['_row'], 'tc_id': tc, 'level': '警告',
                        'category': '实际结果质量',
                        'issue': f'预期风险等级"{exp_final}"与实际"{act_final}"不一致且目标规则未命中',
                    })

        # 4. 实际结果含错误/异常
        for kw in _ERROR_KEYWORDS:
            if kw in actual:
                issues.append({
                    'row': row['_row'], 'tc_id': tc, 'level': '严重',
                    'category': '实际结果质量',
                    'issue': f'实际结果含错误关键词"{kw}"',
                })
                break

    return issues


# ============================================================
# 汇总统计
# ============================================================

def compute_summary(all_issues, total_rows):
    return {
        'total_cases': total_rows,
        'total_issues': len(all_issues),
        'by_level': Counter(i['level'] for i in all_issues),
        'by_category': Counter(i['category'] for i in all_issues),
        'affected_rows': len(set(i['row'] for i in all_issues if i['row'] > 0)),
        'check_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }


# ============================================================
# Excel 标注输出
# ============================================================

def write_annotated_excel(wb, all_issues, summary, output_path, col_map):
    # --- 原始Sheet行级标注 ---
    for ws in wb.worksheets:
        if ws.title == '质量检查':
            continue
        issue_col = ws.max_column + 1

        header_cell = ws.cell(row=1, column=issue_col, value='质量问题')
        header_cell.font = Font(name='微软雅黑', size=9, bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='C0504D', end_color='C0504D', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        row_issues = defaultdict(list)
        for issue in all_issues:
            if issue['row'] > 0:
                row_issues[issue['row']].append(issue)

        for row_num, issues in row_issues.items():
            issue_text = '; '.join(f'[{i["level"]}] {i["issue"]}' for i in issues)
            cell = ws.cell(row=row_num, column=issue_col, value=issue_text)
            cell.font = Font(name='微软雅黑', size=8)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

            max_level = '提示'
            for i in issues:
                if i['level'] == '严重':
                    max_level = '严重'
                    break
                elif i['level'] == '警告' and max_level != '严重':
                    max_level = '警告'

            fills = {
                '严重': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
                '警告': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
                '提示': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            }
            cell.fill = fills[max_level]

        ws.column_dimensions[get_column_letter(issue_col)].width = 60

    # --- "质量检查"汇总Sheet ---
    ws_check = wb.create_sheet('质量检查', 0)

    hdr_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 标题
    ws_check.merge_cells('A1:F1')
    c = ws_check['A1']
    c.value = f'测试报告质量检查 — {summary["check_time"]}'
    c.font = Font(name='微软雅黑', size=14, bold=True, color='4472C4')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws_check.row_dimensions[1].height = 30

    # 摘要
    summary_items = [
        ('检查用例总数', summary['total_cases']),
        ('发现问题总数', summary['total_issues']),
        ('涉及问题行数', summary['affected_rows']),
        ('严重', summary['by_level'].get('严重', 0)),
        ('警告', summary['by_level'].get('警告', 0)),
        ('提示', summary['by_level'].get('提示', 0)),
    ]
    for i, (label, val) in enumerate(summary_items):
        r = i + 3
        ws_check.cell(row=r, column=1, value=label).font = Font(name='微软雅黑', size=9, bold=True)
        ws_check.cell(row=r, column=2, value=val).font = Font(name='微软雅黑', size=9)

    # 按类别统计
    r = len(summary_items) + 4
    ws_check.cell(row=r, column=1, value='按检查维度统计').font = Font(name='微软雅黑', size=10, bold=True)
    r += 1
    for cat, count in summary['by_category'].most_common():
        ws_check.cell(row=r, column=1, value=cat).font = Font(name='微软雅黑', size=9)
        ws_check.cell(row=r, column=2, value=count).font = Font(name='微软雅黑', size=9)
        r += 1

    # 问题明细表
    r += 2
    ws_check.cell(row=r, column=1, value='问题明细').font = Font(name='微软雅黑', size=10, bold=True)
    r += 1

    detail_headers = ['序号', '行号', '用例编号', '严重级别', '检查维度', '问题描述']
    detail_widths = [6, 8, 20, 10, 14, 70]
    for ci, (header, width) in enumerate(zip(detail_headers, detail_widths), 1):
        cell = ws_check.cell(row=r, column=ci, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin
        ws_check.column_dimensions[get_column_letter(ci)].width = width

    r += 1
    level_order = {'严重': 0, '警告': 1, '提示': 2}
    sorted_issues = sorted(all_issues, key=lambda x: (level_order.get(x['level'], 9), x['category']))

    level_fills = {
        '严重': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        '警告': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        '提示': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    }

    for idx, issue in enumerate(sorted_issues, 1):
        row_data = [
            idx,
            issue['row'] if issue['row'] > 0 else '-',
            issue['tc_id'],
            issue['level'],
            issue['category'],
            issue['issue'],
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws_check.cell(row=r, column=ci, value=val)
            cell.font = Font(name='微软雅黑', size=8)
            cell.alignment = Alignment(vertical='top', wrap_text=(ci == 6))
            cell.border = thin
            if ci == 4:
                cell.fill = level_fills.get(str(val), PatternFill())
                cell.alignment = Alignment(horizontal='center', vertical='top')
        r += 1

    ws_check.freeze_panes = 'A2'
    wb.save(output_path)


# ============================================================
# 主检查函数
# ============================================================

def check_report(input_path, output_path=None):
    """
    对测试报告执行全维度质量检查。

    Args:
        input_path: 输入的 Excel 报告路径
        output_path: 输出标注后的 Excel 路径（默认: 原文件名_checked.xlsx）

    Returns:
        dict: 包含统计摘要和 issues 列表
    """
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"文件不存在: {input_path}")

    if output_path is None:
        base, ext = os.path.splitext(input_path)
        output_path = f'{base}_checked{ext}'

    wb = openpyxl.load_workbook(input_path)

    # 扫描所有Sheet，找到可识别格式的数据Sheet
    ws = None
    fmt = 'unknown'
    col_map = {}
    for sheet_name in wb.sheetnames:
        candidate = wb[sheet_name]
        candidate_fmt, candidate_map = detect_format(candidate)
        if candidate_fmt != 'unknown':
            ws = candidate
            fmt = candidate_fmt
            col_map = candidate_map
            break

    if ws is None:
        raise ValueError(f"无法识别的报告格式，已检查 {len(wb.sheetnames)} 个Sheet，均未匹配标准表头")

    print(f'检测报告格式: {fmt} (Sheet: {ws.title})', file=sys.stderr)
    print(f'数据行数: {ws.max_row - 1}', file=sys.stderr)

    rows = read_rows(ws, col_map)
    filename = os.path.basename(input_path)

    print('维度一：判定一致性...', file=sys.stderr)
    issues_1 = check_verdict_consistency(rows)

    print('维度二：数据完整性...', file=sys.stderr)
    issues_2 = check_data_integrity(rows, ws, filename)

    print('维度三：语义重复检测...', file=sys.stderr)
    issues_3 = check_semantic_duplicates(rows)

    print('维度四：实际结果质量...', file=sys.stderr)
    issues_4 = check_result_quality(rows)

    all_issues = issues_1 + issues_2 + issues_3 + issues_4
    summary = compute_summary(all_issues, len(rows))

    write_annotated_excel(wb, all_issues, summary, output_path, col_map)

    summary['output_path'] = output_path
    summary['issues'] = all_issues
    return summary


# ============================================================
# CLI
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description='天策策略测试报告自动化质量检查工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python3 check_report.py 测试报告.xlsx
  python3 check_report.py 测试报告.xlsx -o 检查结果.xlsx
  python3 check_report.py 测试报告.xlsx --json
        """
    )
    parser.add_argument('input', help='输入的 Excel 测试报告文件路径')
    parser.add_argument('-o', '--output', help='输出文件路径（默认: 原文件名_checked.xlsx）')
    parser.add_argument('--json', action='store_true', help='以 JSON 格式输出检查结果摘要')
    parser.add_argument('-v', '--version', action='version', version=f'tiance-report-checker v{VERSION}')

    args = parser.parse_args()

    try:
        result = check_report(args.input, args.output)
    except Exception as e:
        print(f'检查失败: {e}', file=sys.stderr)
        sys.exit(1)

    if args.json:
        json_out = {k: v for k, v in result.items() if k != 'issues'}
        print(json.dumps(json_out, ensure_ascii=False, indent=2))
    else:
        print(f'\n{"=" * 50}')
        print(f'  质量检查完成 — tiance-report-checker v{VERSION}')
        print(f'{"=" * 50}')
        print(f'  检查用例总数:  {result["total_cases"]}')
        print(f'  发现问题总数:  {result["total_issues"]}')
        print(f'  涉及问题行数:  {result["affected_rows"]}')
        print(f'  ─────────────────────────')
        for level in ['严重', '警告', '提示']:
            count = result['by_level'].get(level, 0)
            marker = '!' if level == '严重' and count > 0 else ' '
            print(f'  {marker} {level}: {count}')
        print(f'  ─────────────────────────')
        for cat, count in result['by_category'].most_common():
            print(f'  {cat}: {count}')
        print(f'\n  标注报告已保存至: {result["output_path"]}')


if __name__ == '__main__':
    main()
