#!/usr/bin/env python3
"""
天策策略测试结果 → Excel报告更新脚本（合并版 v3 — 配置驱动比对）

适配15列/16列报告结构，支持颜色标注。

支持五种操作模式：
  1. ID模式（默认）：按用例ID正则匹配测试数据列中的用例标识
     python3 update_report.py results.json report.xlsx [--version V3]

  2. 行号模式：按指定行号更新
     python3 update_report.py --by-row --row 11 --result "绿灯" \
       --token "xxx" --excel report.xlsx

  3. 批量行号模式：从JSON文件批量按行号更新
     python3 update_report.py --by-row --batch rows.json --excel report.xlsx

  4. 概览模式：查看所有行的状态
     python3 update_report.py --overview report.xlsx

  5. 重新比对模式：按比对规则重新标记所有行的测试状态
     python3 update_report.py --relabel report.xlsx [--config jzd|jt|wjfl|bhjc|策略配置JSON路径]

  6. 配置验证模式：校验策略配置JSON文件的格式正确性
     python3 update_report.py --validate-config jzd  (或路径)

results.json 格式（ID模式）:
[
  {
    "id": "CASE_001",
    "expected": "绿灯",
    "batch": 1,
    "version": "V3",
    "rs": "2",
    "dt": "绿灯",
    "dtCode": "greenalert",
    "uuid": "xxx",
    "token": "xxx",
    "err": "",
    "pass": true
  }
]

rows.json 格式（行号模式）:
[
  {"row": 11, "result": "绿灯", "token": "...", "time": "...", "status": "已完成"},
  {"row": 12, "result": "黄灯", "token": "...", "time": "...", "status": "已完成"}
]

策略配置JSON格式 (strategies/{policyCode}.json):
{
  "policyCode": "jzd",
  "policyName": "集中度策略",
  "comparisonRules": [
    {
      "type": "alias_group",
      "trigger": ["不触发", "通过", "Accept", "全部通过"],
      "aliases": [["Accept", "通过", "不触发"]],
      "description": "预期不触发时，实际结果需包含同组别名"
    },
    {
      "type": "keyword_present",
      "trigger": ["触发", "超限"],
      "target": "超限",
      "description": "预期触发时，实际结果需包含'超限'"
    },
    {
      "type": "keyword_absent",
      "trigger": ["不触发", "无超限"],
      "target": "超限",
      "description": "预期不触发时，实际结果不应包含'超限'"
    },
    {
      "type": "numeric_tolerance",
      "trigger": null,
      "expectedPattern": "评分[=≈]?\\\\s*([\\\\d.]+)",
      "actualPattern": "风险预警总分[=≈]?\\\\s*([\\\\d.]+)",
      "tolerance": 1.0,
      "description": "数值在容差范围内即通过"
    },
    {
      "type": "exploratory",
      "trigger": ["待探索"],
      "indicators": ["决策结果", "低风险", "中低风险", "中高风险", "高风险"],
      "description": "探索性用例，实际结果包含有效指示器即通过"
    },
    {
      "type": "regex_match",
      "triggerPattern": ".*低风险.*",
      "actualPattern": ".*低风险.*",
      "description": "正则匹配"
    },
    {
      "type": "pass_through",
      "trigger": ["Accept"],
      "matchActual": ["Accept", "通过"],
      "description": "预期包含trigger时，实际需包含matchActual中任一项"
    },
    {
      "type": "invalid_expectation",
      "trigger": ["次级", "可疑", "损失"],
      "description": "此预期结果不应出现，直接标记未通过"
    }
  ]
}
"""
import json, re, sys, argparse
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("需要 openpyxl: pip install openpyxl")
    sys.exit(1)

# ── 从 comparison_engine 导入比对相关常量和函数 ──
from comparison_engine import (
    load_strategy_config, validate_strategy_config,
    compare_generic, resolve_comparator,
)

# ── 从 excel_formatter 导入格式化相关常量和函数 ──
from excel_formatter import (
    backup_excel, get_col_map, detect_col_format, get_uuid_col,
    color_status_cell,
)


# ═══════════════════════════════════════════════════════
#  各模式核心逻辑
# ═══════════════════════════════════════════════════════

def show_overview(ws, col):
    """显示所有行的状态概览"""
    scene_col = col.get("测试场景", 5)
    expected_col = col.get("预期结果", 9)
    actual_col = col.get("实际结果", 10)
    status_col = col.get("测试状态", 11)
    token_col = get_uuid_col(col)

    print(f"{'行号':>4}  {'测试场景':<30}  {'预期':<15}  {'实际':<15}  {'状态':<6}  {'UUID/Token':<20}")
    print("-" * 100)
    for row in range(2, ws.max_row + 1):
        scenario = ws.cell(row=row, column=scene_col).value or ""
        expected = ws.cell(row=row, column=expected_col).value or ""
        actual = ws.cell(row=row, column=actual_col).value or ""
        exec_st = ws.cell(row=row, column=status_col).value or ""
        token = ws.cell(row=row, column=token_col).value or ""
        if scenario or expected:
            print(
                f"{row:>4}  {str(scenario)[:30]:<30}  "
                f"{str(expected)[:15]:<15}  {str(actual)[:15]:<15}  "
                f"{str(exec_st)[:6]:<6}  {str(token)[:20]:<20}"
            )


def relabel_sheet(wb, ws, col, comparator=None):
    """对比对预期/实际结果，重新标记测试状态并着色"""
    expected_col = col.get("预期结果", 9)
    actual_col = col.get("实际结果", 10)
    status_col = col.get("测试状态", 11)

    # 如果没指定 comparator，尝试从 sheet 名自动推断
    if not comparator:
        comparator = resolve_comparator(sheet_name=ws.title)

    if not comparator:
        print(f"  {ws.title}: 未找到比对规则，跳过")
        return 0, 0

    changes = {'通过': 0, '未通过': 0}
    for r in range(2, ws.max_row + 1):
        exp = ws.cell(row=r, column=expected_col).value or ''
        act = ws.cell(row=r, column=actual_col).value or ''
        if not exp and not act:
            continue

        result = comparator(exp, act)
        if result:
            ws.cell(row=r, column=status_col).value = result
            color_status_cell(ws, r, status_col, result)
            changes[result] = changes.get(result, 0) + 1

    print(f"  {ws.title}: 通过={changes['通过']} 未通过={changes['未通过']}")
    return changes['通过'], changes['未通过']


def update_row_by_result(ws, row, r, col, ver):
    """用ID模式的结果字典更新一行（适配15/16列结构）"""
    passed = r.get("pass", False)
    dt = r.get("dt", "")
    dtCode = r.get("dtCode", "")
    err = r.get("err", "")
    uuid_val = r.get("uuid", "")
    token = r.get("token", "")
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 实际结果（列10）
    if "实际结果" in col:
        if passed:
            val = f"决策成功\n决策结果={dt}({dtCode})\n[{ver}运行区]"
        else:
            val = f"决策流调用失败\n{err}\n[{ver}运行区]"
        ws.cell(row=row, column=col["实际结果"]).value = val

    # 测试状态（列11）
    status = "通过" if passed else "未通过"
    if "测试状态" in col:
        ws.cell(row=row, column=col["测试状态"]).value = status
        color_status_cell(ws, row, col["测试状态"], status)

    # 测试时间（列12）
    if "测试时间" in col:
        ws.cell(row=row, column=col["测试时间"]).value = now_str

    # 测试流水号UUID（列13，统一列名）
    uuid_col = get_uuid_col(col)
    if uuid_col:
        ws.cell(row=row, column=uuid_col).value = (
            f"{uuid_val}" if uuid_val else f"[{ver}]{token}"
        )


def update_row_simple(ws, row, result=None, token=None, time_str=None,
                      status=None, notes=None, col=None):
    """行号模式的简单更新（适配15/16列结构）"""
    actual_col = col.get("实际结果", 10) if col else 10
    uuid_col = get_uuid_col(col) if col else 13
    time_col = col.get("测试时间", 12) if col else 12
    status_col = col.get("测试状态", 11) if col else 11
    notes_col = col.get("备注") if col else None

    if result:
        ws.cell(row=row, column=actual_col).value = result
    if token:
        ws.cell(row=row, column=uuid_col).value = token
    if time_str:
        ws.cell(row=row, column=time_col).value = time_str

    status_val = status or ("通过" if result else None)
    if status_val and status_col:
        ws.cell(row=row, column=status_col).value = status_val
        color_status_cell(ws, row, status_col, status_val)

    # 写入备注（16列格式时可用）
    if notes and notes_col:
        ws.cell(row=row, column=notes_col).value = notes


# ═══════════════════════════════════════════════════════
#  实际覆盖率计算 & diff（v2.4.2 Phase 3-2）
# ═══════════════════════════════════════════════════════

def compute_coverage_actual(results, ws, col, report_path):
    """从测试结果和 Excel 工作表计算实际覆盖率，写入 coverage_actual.json。
    若同目录有 coverage_pre_report.json，同时输出 coverage_diff.json。
    """
    result_map = {r["id"]: r for r in results}

    # 从 Excel 读取每条用例的模块和预期结果
    module_col = col.get("模块", col.get("用例类型", 2))
    expected_col = col.get("预期结果", 9)
    actual_col = col.get("实际结果", 10)
    status_col = col.get("测试状态", 11)
    id_col_num = col.get("测试数据", 14)
    id_pattern = re.compile(r'TC_\d{3}')

    case_data = []  # (id, module, expected, actual, status, result)
    for row in range(2, ws.max_row + 1):
        td = str(ws.cell(row=row, column=id_col_num).value or "")
        m = id_pattern.search(td)
        if not m:
            continue
        tc_id = m.group(0)
        r = result_map.get(tc_id)
        module = str(ws.cell(row=row, column=module_col).value or "")
        expected = str(ws.cell(row=row, column=expected_col).value or "")
        actual = str(ws.cell(row=row, column=actual_col).value or "")
        status = str(ws.cell(row=row, column=status_col).value or "")
        case_data.append((tc_id, module, expected, actual, status, r))

    total = len(case_data)
    passed = sum(1 for _, _, _, _, s, _ in case_data if s in ("通过", "pass", "PASS"))
    failed = sum(1 for _, _, _, _, s, _ in case_data if s in ("未通过", "fail", "FAIL"))
    errored = sum(1 for _, _, _, _, _, r in case_data if r and r.get("err"))
    executed = sum(1 for _, _, _, _, _, r in case_data if r and r.get("rs") == "2")

    # 按模块分组
    by_module = {}
    for tc_id, module, exp, act, status, r in case_data:
        by_module.setdefault(module, {"total": 0, "passed": 0, "failed": 0, "errored": 0})
        by_module[module]["total"] += 1
        if status in ("通过", "pass", "PASS"):
            by_module[module]["passed"] += 1
        elif status in ("未通过", "fail", "FAIL"):
            by_module[module]["failed"] += 1
        if r and r.get("err"):
            by_module[module]["errored"] += 1

    coverage = {
        "totalCases": total,
        "executed": executed,
        "passed": passed,
        "failed": failed,
        "errored": errored,
        "passRate": round(passed / max(total, 1) * 100, 1),
        "executionRate": round(executed / max(total, 1) * 100, 1),
        "byModule": by_module,
    }

    # 写入 coverage_actual.json
    actual_path = Path(report_path).parent / "coverage_actual.json"
    actual_path.write_text(json.dumps(coverage, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"实际覆盖率报告: {actual_path}", file=sys.stderr)

    # 与预期覆盖率 diff
    pre_path = Path(report_path).parent / "coverage_pre_report.json"
    if pre_path.exists():
        pre = json.loads(pre_path.read_text(encoding="utf-8"))
        diff = _compute_coverage_diff(pre, coverage)
        diff_path = Path(report_path).parent / "coverage_diff.json"
        diff_path.write_text(json.dumps(diff, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"覆盖率差异报告: {diff_path}", file=sys.stderr)


def _compute_coverage_diff(pre_report, actual):
    """比较预期覆盖率与实际覆盖率，输出差异摘要。"""
    diff = {"summary": {}, "gaps": []}

    pre_total = pre_report.get("totalCases", 0)
    act_total = actual.get("totalCases", 0)
    diff["summary"] = {
        "expectedCases": pre_total,
        "actualCases": act_total,
        "executionRate": actual.get("executionRate", 0),
        "passRate": actual.get("passRate", 0),
    }

    # 规则维度 diff
    pre_rules = pre_report.get("dimensions", {}).get("rules", {}).get("detail", {})
    act_modules = actual.get("byModule", {})
    for rule_code, rule_info in pre_rules.items():
        rule_name = rule_info.get("name", rule_code)
        # 在 actual 中查找对应模块的用例通过情况
        # 规则用例的 module 通常是 "规则"
        rule_module = act_modules.get("规则", {})
        total_in_module = rule_module.get("total", 0)
        passed_in_module = rule_module.get("passed", 0)

    # 模块级 diff
    pre_by_module = pre_report.get("dimensions", {})
    for dim_name, dim_data in pre_by_module.items():
        if isinstance(dim_data, dict) and "total" in dim_data:
            pre_count = dim_data.get("total", 0)
            # 尝试在实际中找到对应
            for act_mod, act_data in act_modules.items():
                if dim_name.lower() in act_mod.lower() or act_mod in dim_name:
                    act_passed = act_data.get("passed", 0)
                    if act_passed < pre_count:
                        diff["gaps"].append({
                            "dimension": dim_name,
                            "expected": pre_count,
                            "actualPassed": act_passed,
                            "gap": pre_count - act_passed,
                        })

    return diff


# ═══════════════════════════════════════════════════════
#  主入口
# ═══════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="天策测试结果 Excel 更新（v3 配置驱动比对）"
    )
    # 位置参数（ID模式）
    parser.add_argument("results", nargs="?", help="结果JSON文件路径（ID模式）")
    parser.add_argument("report", nargs="?", help="Excel报告文件路径")

    # 模式选择
    parser.add_argument("--by-row", action="store_true",
                        help="行号模式（配合 --row 或 --batch）")
    parser.add_argument("--overview", action="store_true",
                        help="概览模式：查看所有行状态")
    parser.add_argument("--relabel", action="store_true",
                        help="重新比对模式：按比对规则重新标记测试状态")
    parser.add_argument("--validate-config", metavar="CONFIG",
                        help="验证策略配置JSON文件（传 policyCode 或文件路径）")

    # ID模式选项
    parser.add_argument("--version", default="V1", help="版本标签（默认V1）")
    parser.add_argument("--id-col", default="测试数据", help="用于匹配用例ID的列名")
    parser.add_argument("--id-pattern",
                        default=r"(PRE_ALL_\d+|PRE_R\d+_\d+|POST_ALL_\d+|POST_R\d+_\d+|WJFL_V\d+_\d+|JTFUNC_\d+|JTDEC_\d+|CASE_\d+|TC_\d+|RULE_\w+)",
                        help="用例ID正则")

    # 行号模式选项
    parser.add_argument("--excel", help="Excel文件路径（行号模式）")
    parser.add_argument("--sheet", help="Sheet名称（行号模式，默认active）")
    parser.add_argument("--row", type=int, help="要更新的行号")
    parser.add_argument("--result", help="实际执行结果")
    parser.add_argument("--token", help="测试Token/流水号")
    parser.add_argument("--time", help="执行时间")
    parser.add_argument("--status", default="已完成", help="运行状态")
    parser.add_argument("--notes", help="备注")
    parser.add_argument("--batch", help="批量更新JSON文件路径（行号模式）")

    # 比对配置选项（--config 为新版，--strategy 为旧版兼容）
    parser.add_argument("--config", default=None,
                        help="策略配置：policyCode（如 jzd）或 JSON 文件路径")
    parser.add_argument("--strategy", choices=['jzd', 'jt', 'wjfl', 'bhjc'],
                        help="[已弃用] 指定硬编码比对函数（建议使用 --config）")

    # 通用选项
    parser.add_argument("--no-backup", action="store_true", help="跳过备份")

    args = parser.parse_args()

    # ── 配置验证模式 ──
    if args.validate_config:
        print(f"验证策略配置: {args.validate_config}")
        is_valid, errors, warnings = validate_strategy_config(args.validate_config)

        if warnings:
            print("\n⚠ 警告:")
            for w in warnings:
                print(f"  - {w}")

        if is_valid:
            print("\n✓ 配置验证通过，格式正确。")
            # 额外：尝试加载并显示摘要
            try:
                rules, config_path = load_strategy_config(args.validate_config)
                print(f"  配置文件: {config_path}")
                print(f"  规则数量: {len(rules)}")
                for i, rule in enumerate(rules):
                    desc = rule.get('description', '(无描述)')
                    print(f"    [{i}] {rule.get('type', '?')}: {desc}")
            except Exception:
                pass
            return
        else:
            print("\n✗ 配置验证失败:")
            for e in errors:
                print(f"  - {e}")
            sys.exit(1)

    # ── 概览模式 ──
    if args.overview:
        report_path = args.report or args.excel
        if not report_path:
            print("请指定Excel文件路径", file=sys.stderr)
            sys.exit(1)
        wb = load_workbook(report_path)
        ws = wb.active
        col = get_col_map(ws)
        fmt = detect_col_format(col)
        print(f"[{fmt}] {report_path}")
        show_overview(ws, col)
        return

    # ── 重新比对模式 ──
    if args.relabel:
        report_path = args.report or args.excel
        if not report_path:
            print("请指定Excel文件路径", file=sys.stderr)
            sys.exit(1)
        if not args.no_backup:
            backup_excel(report_path)
        wb = load_workbook(report_path)

        # 解析比对器（如果指定了 --config 或 --strategy，先为 active sheet 解析一次）
        pre_comparator = None
        if args.config or args.strategy:
            pre_comparator = resolve_comparator(
                config_arg=args.config,
                strategy_arg=args.strategy,
            )

        print("重新比对中...")
        total_pass = 0
        total_fail = 0
        if pre_comparator:
            # 指定了比对器 → 仅处理 active sheet
            ws = wb.active
            col = get_col_map(ws)
            p, f = relabel_sheet(wb, ws, col, pre_comparator)
            total_pass += p
            total_fail += f
        else:
            # 未指定 → 遍历所有 sheet，逐个自动推断
            for sn in wb.sheetnames:
                ws = wb[sn]
                col = get_col_map(ws)
                p, f = relabel_sheet(wb, ws, col)
                total_pass += p
                total_fail += f

        wb.save(report_path)
        print(f"\n总计: 通过={total_pass} 未通过={total_fail}")
        return

    # ── 行号模式 ──
    if args.by_row:
        excel_path = args.excel or args.report
        if not excel_path:
            print("请指定 --excel 文件路径", file=sys.stderr)
            sys.exit(1)

        if not args.no_backup:
            backup_excel(excel_path)

        wb = load_workbook(excel_path)
        if args.sheet and args.sheet in wb.sheetnames:
            ws = wb[args.sheet]
        else:
            ws = wb.active
        col = get_col_map(ws)
        fmt = detect_col_format(col)
        print(f"[{fmt}]")

        if args.batch:
            with open(args.batch, "r", encoding="utf-8") as f:
                items = json.load(f)
            for item in items:
                row = item["row"]
                update_row_simple(
                    ws, row,
                    result=item.get("result"),
                    token=item.get("token"),
                    time_str=item.get("time"),
                    status=item.get("status", "已完成"),
                    notes=item.get("notes"),
                    col=col,
                )
                print(f"Row {row}: 已更新 → {item.get('result', 'N/A')}")
            wb.save(excel_path)
            print(f"已保存: {excel_path} (共更新 {len(items)} 行)")
            return

        if not args.row:
            print("请指定 --row 或 --batch", file=sys.stderr)
            sys.exit(1)

        update_row_simple(
            ws, args.row,
            result=args.result,
            token=args.token,
            time_str=args.time,
            status=args.status,
            notes=args.notes,
            col=col,
        )

        # 如果提供了 --config，在行号模式下也做一次结果比对来判定状态
        if args.config and args.result:
            try:
                rules, _ = load_strategy_config(args.config)
                expected_col = col.get("预期结果", 9)
                status_col = col.get("测试状态", 11)
                exp = ws.cell(row=args.row, column=expected_col).value or ''
                verdict = compare_generic(exp, args.result, rules)
                if verdict:
                    ws.cell(row=args.row, column=status_col).value = verdict
                    color_status_cell(ws, args.row, status_col, verdict)
                    print(f"  配置比对结果: {verdict}")
            except Exception as e:
                print(f"  配置比对跳过: {e}", file=sys.stderr)

        wb.save(excel_path)
        print(f"Row {args.row}: 已更新 → {args.result or 'N/A'}")
        return

    # ── ID模式（默认） ──
    if not args.results or not args.report:
        print("ID模式需要提供 results.json 和 report.xlsx 两个位置参数", file=sys.stderr)
        print("用法: python3 update_report.py results.json report.xlsx [--version V1]",
              file=sys.stderr)
        sys.exit(1)

    with open(args.results) as f:
        results = json.load(f)
    result_map = {r["id"]: r for r in results}

    if not args.no_backup:
        backup_excel(args.report)

    wb = load_workbook(args.report)
    ws = wb.active
    col = get_col_map(ws)
    fmt = detect_col_format(col)
    print(f"[{fmt}]")

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ver = args.version
    id_pattern = re.compile(args.id_pattern)
    id_col_num = col.get(args.id_col, 14)  # 15列结构中测试数据是第14列

    # 如果提供了 --config，在ID模式下用配置比对来覆盖状态判定
    config_comparator = None
    if args.config:
        try:
            rules, config_path = load_strategy_config(args.config)
            print(f"ID模式启用配置比对: {config_path.name}")
            config_comparator = lambda exp, act: compare_generic(exp, act, rules)
        except Exception as e:
            print(f"  配置比对加载失败: {e}，使用 results.json 中的 pass 字段", file=sys.stderr)

    updated = 0
    expected_col = col.get("预期结果", 9)
    actual_col = col.get("实际结果", 10)
    status_col = col.get("测试状态", 11)

    for row in range(2, ws.max_row + 1):
        td = str(ws.cell(row=row, column=id_col_num).value or "")
        m = id_pattern.search(td)
        if not m or m.group(0) not in result_map:
            continue

        r = result_map[m.group(0)]
        update_row_by_result(ws, row, r, col, ver)

        # 配置比对覆盖：用配置规则重新判定状态
        if config_comparator:
            exp = ws.cell(row=row, column=expected_col).value or ''
            act = ws.cell(row=row, column=actual_col).value or ''
            verdict = config_comparator(exp, act)
            if verdict:
                ws.cell(row=row, column=status_col).value = verdict
                color_status_cell(ws, row, status_col, verdict)

        updated += 1

    # 输出实际覆盖率报告 + 与预期覆盖率的 diff
    try:
        compute_coverage_actual(results, ws, col, args.report)
    except Exception as e:
        print(f"覆盖率报告生成跳过: {e}", file=sys.stderr)

    wb.save(args.report)
    pass_n = sum(1 for r in results if r.get("pass"))
    fail_n = len(results) - pass_n
    print(f"已更新 {updated} 行。通过: {pass_n}，失败: {fail_n}")


if __name__ == "__main__":
    main()
