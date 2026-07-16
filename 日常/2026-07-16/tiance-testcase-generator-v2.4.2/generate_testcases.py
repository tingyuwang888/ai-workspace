#!/usr/bin/env python3
"""
天策策略测试用例生成引擎 v2.4.2

从 parsed_strategy.json 生成完整的结构化测试用例。
覆盖：规则级(正/反/边界/条件拆解) + 函数级(含格式兼容/健壮性) +
      决策流分支(含DEFAULT/异常路径) + 预警等级 + 跨子策略综合场景

v2.4.2 变更：
  - _extract_threshold_fallback 增强：_resolve_field() 前缀模糊匹配，修复截断字段名（如"按集中度管理策略限额标准"→"按集中度管理策略限额标准_离线集中度"）
  - has_threshold_condition / is_strict_operator 扩展：识别 gte_field / gt_field / lte_field / lt_field / gte_float / lte_float 类型
  - construct_param_value 精确化：gte_field 与 gt_field 边界值分离（1000000 vs 1000001）
  - field_code 碰撞检测：AND 拆解前检查 field_code 去重，避免 areaprovince 双值覆盖
  - 扁平架构新增三类生成器：generate_ruleset_cross_matrix（2^n 联动矩阵）、generate_region_templates（区域特化模板）、apiRequiredFields 自检注入
  - section 2/3 同步使用 _resolve_field，三个解析分支全部支持模糊匹配

v2.4.1 变更：
  - 新增 detect_strategy_architecture()：根据 ruleSets 的 bizScenarios × enterpriseType 路由维度区分 tripartite / flat 架构
  - 架构感知的生成路由：tripartite 走 generate_flow_coverage_cases / generate_alert_level_cases / generate_combined_scenarios；flat 走 generate_simple_flow_cases / generate_simple_combined_cases
  - 路由参数注入加架构守卫：C_S_BIZSCENARIO / S_S_ENTERPRISETYPE 仅对 tripartite 注入，避免污染扁平架构策略（如 DF_PRE_CONC_001）的用例
  - 修复"串台"问题：保后检查（三子策略）专用模板不再生成到扁平架构策略的用例中

v2.4.0 变更：
  - 新增 --feedback / -f 参数：接收上轮 Agent Loop 的 feedback.json，自动修正用例
  - 支持 fixParams（修正入参）、adjustExpected（调整预期）、removeCases（移除无效用例）
  - 匹配优先级：targetRule > scenario 关键词 > desc 关键词
  - CLI 从 sys.argv 迁移到 argparse，向后兼容位置参数

v2.3.1 变更：
  - BUG-2 fix: Excel 报告 Col5/Col6 数据与表头错位修正（caseType→用例类型，ruleSet/rule→测试场景）
  - BUG-3 fix: 规则级用例自动注入路由参数（C_S_BIZSCENARIO / S_S_ENTERPRISETYPE），基于 ruleSet 元数据推导

v2.3.0 变更：
  - 新增 export_excel：同时输出 JSON + Excel 测试报告（16列、蓝底白字表头、汇总+用例双Sheet）
  - openpyxl 为可选依赖，未安装时自动跳过 Excel 导出

v2.2.0 变更：
  - get_detail_template 重构：if/elif链 → 数据驱动有序规则表(_DETAIL_RULES)
  - 新增时间窗口边界用例：自动检测"距今≤N天"模式，生成日期临界测试
  - 新增 _safe_date_days_ago / _extract_day_window 辅助函数

v2.1.0 变更：
  - 正则模式列表抽为模块级 _field_patterns()，消除三处重复
  - 支持 || OR 运算符（除中文"或"外）
  - 中文全角括号（）自动归一化为半角
  - 枚举值内的"或"不再被误判为 OR（如"为d或D"）
  - thirdPartyMock 输出结构化对象（interfaces/fields/mockHint）
  - slash_enum 完整捕获所有斜杠分隔的枚举值
  - 负数表达式支持（如 = -999）
  - gte_years 闰年安全日期计算
  - SAFE 基线用例填充实际安全参数值
  - generate_all 增加错误处理与友好提示
  - 函数用例增加模糊名称匹配回退
  - 清理 construct_param_value 死代码
"""
import argparse
import json
import re
import sys
from pathlib import Path
from datetime import datetime, timedelta

VERSION = "2.4.1"

# v2.3.1: Excel 报告输出（可选依赖）
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# 公共参数（每条用例都需包含）
# ---------------------------------------------------------------------------
COMMON_PARAMS = {
    "S_S_BIZID": "TC_BIZ_{tc_id}",
    "S_S_CUSTNO": "TC_CUST_{tc_id}",
    "S_S_ORGCODE": "010117",
    "S_E_APIMODEL": "30",
}

# v2.1.0: 常用字段安全默认值（用于 SAFE 基线用例）
SAFE_DEFAULTS = {
    "C_N_LEGREPPERSONCHANGECOUNT_": 0,
    "C_N_SHAREHOLDERCHANGECOUNT_": 0,
    "C_N_BIZSCOPECHANGECOUNT_": 0,
    "C_N_REGISTCAPIDECOUNT_": 0,
    "C_N_REGISTCAPIDECAMOUNT_": 0,
    "C_N_REGISTCAPIINCCOUNT_": 0,
    "C_N_REGISTCAPIINCAMOUNT_": 0,
    "C_N_DISHONESTCOUNT_": 0,
    "C_N_EXECUTEDCOUNT_": 0,
    "C_N_HIGHCONSUMECOUNT_": 0,
    "C_N_TERMINATEDCOUNT_": 0,
    "C_N_PRESERVATIONCOUNT_": 0,
    "C_N_DEFENDANTCASECOUNT_": 0,
    "C_N_LOANDISPUTEDOCCOUNT_": 0,
    "C_N_BANKRUPTCOUNT_": 0,
    "C_N_LIQUIDATIONCOUNT_": 0,
    "C_N_TAXARREARSCOUNT_": 0,
    "C_N_TAXVIOLATIONCOUNT_": 0,
    "C_N_EQUITYFREEZECOUNT_": 0,
    "C_N_EQUITYPLEDGECOUNT_": 0,
    "C_N_SERIOUSVIOLATIONCOUNT_": 0,
    "C_N_JUDICIALAUCTIONCOUNT_": 0,
    "C_N_ADMINPENALTYCOUNT_": 0,
    "C_N_BIZABNORMALCOUNT_": 0,
    "C_N_OVERDUELOANCOUNT_": 0,
    "C_N_ATTENTIONLOANCOUNT_": 0,
    "C_N_OVERDUEGUARANTEECOUNT_": 0,
    "C_N_SELFPROSECUTIONCOUNT_": 0,
    "C_N_NEGATIVEPUBLICOPINIONCOUNT_": 0,
    "C_N_COURTANNOUNCEMENTCOUNT_": 0,
    "C_N_COURTNOTICECOUNT_": 0,
}

# ---------------------------------------------------------------------------
# 详情字段模板（按业务含义分类）
# ---------------------------------------------------------------------------
RECENT_DATE = (datetime.now() - timedelta(days=15)).strftime("%Y-%m-%d")
RECENT_DATE_2 = (datetime.now() - timedelta(days=45)).strftime("%Y-%m-%d")

DETAIL_TEMPLATES = {
    "change_record": [
        {"ChangeDate": RECENT_DATE, "ProjectName": "{keyword}", "BeforeList": "变更前内容", "AfterList": "变更后内容"}
    ],
    "revoke_record": [
        {"CancelDate": RECENT_DATE, "RevokeDate": "", "RevokeReason": "决议解散"}
    ],
    "dishonest_record": [
        {"Liandate": RECENT_DATE, "Anno": "未履行法院判决", "ExecuteGov": "某市人民法院", "ExecuteStatus": "执行中", "Publicdate": RECENT_DATE}
    ],
    "executed_record": [
        {"Liandate": RECENT_DATE, "Anno": "民事判决", "ExecuteGov": "某市人民法院", "Biaodi": 100000, "SuspectedApplicant": "测试企业"}
    ],
    "high_consume_record": [
        {"Liandate": RECENT_DATE, "Anno": "限制消费令", "ExecuteGov": "某市人民法院", "CaseCode": "(2026)某执限字第001号"}
    ],
    "preservation_record": [
        {"Liandate": RECENT_DATE, "Anno": "财产保全", "ExecuteGov": "某市人民法院", "Amount": 500000}
    ],
    "terminated_record": [
        {"Liandate": RECENT_DATE, "Anno": "终结本次执行", "ExecuteGov": "某市人民法院", "ExecuteAmount": 200000}
    ],
    "defendant_case": [
        {"CaseRole": "被告", "CaseAmount": 6000000, "CaseType": "借款合同纠纷", "FilingDate": RECENT_DATE}
    ],
    "loan_dispute_doc": [
        {"CaseRole": "被告", "CaseType": "借款合同纠纷", "Court": "某市人民法院", "FilingDate": RECENT_DATE}
    ],
    "court_announcement": [
        {"CaseRole": "被告", "CaseType": "金融借款合同纠纷", "Court": "某市人民法院", "AnnouncementDate": RECENT_DATE}
    ],
    "court_notice": [
        {"CaseRole": "被告", "CaseType": "票据纠纷", "Court": "某市人民法院", "NoticeDate": RECENT_DATE}
    ],
    "bankrupt_record": [
        {"CompanyName": "测试企业", "FilingDate": RECENT_DATE, "Court": "某市人民法院", "CaseType": "破产清算"}
    ],
    "tax_arrears_record": [
        {"TaxType": "增值税", "ArrearsAmount": 50000, "TaxOrg": "某税务局", "Period": "2026年1月"}
    ],
    "tax_violation_record": [
        {"ViolationType": "偷税", "PenaltyAmount": 100000, "DecisionDate": RECENT_DATE, "TaxOrg": "某税务局"}
    ],
    "equity_freeze_record": [
        {"FrozenAmount": 1000000, "FreezeCourt": "某市人民法院", "FreezeDate": RECENT_DATE, "FreezePeriod": "2年"}
    ],
    "equity_pledge_record": [
        {"PledgeAmount": 500000, "Pledgor": "测试企业", "Pledgee": "某银行", "RegistrationDate": RECENT_DATE}
    ],
    "serious_violation_record": [
        {"ViolationType": "严重违法", "DecisionDate": RECENT_DATE, "DecisionOffice": "某市场监管局", "RemoveDate": ""}
    ],
    "judicial_auction_record": [
        {"AuctionDate": RECENT_DATE, "AuctionAmount": 2000000, "Court": "某市人民法院", "Subject": "不动产"}
    ],
    "admin_penalty_record": [
        {"PenaltyDate": RECENT_DATE, "PenaltyAmount": 50000, "PenaltyReason": "违规经营", "DecisionOffice": "某市场监管局"}
    ],
    "env_penalty_record": [
        {"PenaltyDate": RECENT_DATE, "PenaltyAmount": 30000, "PenaltyReason": "超标排放", "DecisionOffice": "某生态环境局"}
    ],
    "biz_abnormal_record": [
        {"AddReason": "通过登记住所无法联系", "AddDate": RECENT_DATE, "RemoveReason": "", "RemoveDate": "", "DecisionOffice": "某市场监管局"}
    ],
    "overdue_loan_record": [
        {"AccountNo": "LOAN001", "OverdueDays": 90, "OverdueAmount": 500000, "Lender": "某银行"}
    ],
    "attention_loan_record": [
        {"AccountNo": "LOAN002", "Status": "关注", "Balance": 300000, "Lender": "某银行"}
    ],
    "overdue_guarantee_record": [
        {"GuaranteeNo": "GUA001", "OverdueDays": 60, "OverdueAmount": 200000, "Borrower": "被担保企业"}
    ],
    "attention_guarantee_record": [
        {"GuaranteeNo": "GUA002", "Status": "关注", "Balance": 150000, "Borrower": "被担保企业"}
    ],
    "personal_dishonest_record": [
        {"Liandate": RECENT_DATE, "Anno": "未履行法院判决", "ExecuteGov": "某市人民法院", "ExecuteStatus": "执行中"}
    ],
    "personal_executed_record": [
        {"Liandate": RECENT_DATE, "Anno": "民事判决", "ExecuteGov": "某市人民法院", "Biaodi": 80000}
    ],
    "personal_high_consume_record": [
        {"Liandate": RECENT_DATE, "Anno": "限制消费令", "ExecuteGov": "某市人民法院"}
    ],
    "personal_terminated_record": [
        {"Liandate": RECENT_DATE, "Anno": "终结本次执行", "ExecuteGov": "某市人民法院"}
    ],
    "related_revoke_record": [
        {"RelatedCorpName": "关联测试企业", "CancelDate": RECENT_DATE, "RevokeDate": "", "RevokeReason": "决议解散"}
    ],
    "shareholder_dishonest_record": [
        {"CompanyName": "股东公司A", "KeyNo": "KEY001", "Liandate": RECENT_DATE, "Anno": "未履行判决", "ExecuteGov": "某市人民法院"}
    ],
    "investee_dishonest_record": [
        {"CompanyName": "被投资公司B", "KeyNo": "KEY002", "Liandate": RECENT_DATE, "Anno": "合同纠纷", "ExecuteGov": "某市人民法院"}
    ],
    "self_prosecution_case": [
        {"CaseRole": "原告", "CaseAmount": 12000000, "CaseType": "合同纠纷", "FilingDate": RECENT_DATE}
    ],
    "negative_news": [
        {"Title": "某企业涉嫌违规", "Source": "某媒体", "PublishDate": RECENT_DATE, "Sentiment": "负面"}
    ],
    "qualification_cert": [
        {"Name": "建筑工程施工总承包一级", "EndDate": "2027-06-01", "Status": "正常"}
    ],
    "qualification_near_expiry": [
        {"Name": "建筑工程施工总承包一级", "EndDate": "", "Status": "正常"}
    ],
    # v2.2.0: 原 get_detail_template 内联模板提升为正式条目
    "legal_rep_change": [
        {"ChangeDate": RECENT_DATE, "ProjectName": "法定代表人变更", "BeforeList": "张三", "AfterList": "李四"}
    ],
    "shareholder_change": [
        {"ChangeDate": RECENT_DATE, "ProjectName": "股东变更", "BeforeList": "张三", "AfterList": "李四"}
    ],
    "biz_scope_change": [
        {"ChangeDate": RECENT_DATE, "ProjectName": "经营范围变更", "BeforeList": "技术服务", "AfterList": "技术咨询"}
    ],
    "regist_capi_decrease": [
        {"ChangeDate": RECENT_DATE, "ProjectName": "注册资本变更", "BeforeList": "5000万元", "AfterList": "2000万元"}
    ],
    "regist_capi_increase": [
        {"ChangeDate": RECENT_DATE, "ProjectName": "注册资本变更", "BeforeList": "2000万元", "AfterList": "5000万元"}
    ],
    "bid_record": [
        {"ProjectName": "某市政工程项目", "BidAmount": 5000000, "WinDate": RECENT_DATE}
    ],
    "generic_detail": [
        {"detail": "测试详情数据", "date": RECENT_DATE}
    ],
}


def get_detail_template(field_name):
    """根据字段中文名选择合适的详情模板。
    v2.2.0: 数据驱动有序规则表，新增匹配只需在 _DETAIL_RULES 中追加一行。"""
    for keywords, template_key in _DETAIL_RULES:
        if all(kw in field_name for kw in keywords):
            return json.dumps(DETAIL_TEMPLATES[template_key], ensure_ascii=False)
    return json.dumps(DETAIL_TEMPLATES["generic_detail"], ensure_ascii=False)


# v2.2.0: 有序规则表 — 每条 (关键词元组, DETAIL_TEMPLATES键)
# 匹配逻辑：all(keywords) 必须全部出现在字段名中，按顺序匹配首条命中。
_DETAIL_RULES = [
    # ── 工商变更类（含特化 ProjectName）──
    (("法定代表人变更",),                        "legal_rep_change"),
    (("股权变更",),                              "shareholder_change"),
    (("经营范围变更",),                          "biz_scope_change"),
    (("注册资本减少", "详情"),                    "regist_capi_decrease"),
    (("注册资本增加", "详情"),                    "regist_capi_increase"),
    # ── 注销吊销类 ──
    (("关联", "注销吊销"),                       "related_revoke_record"),
    (("注销吊销",),                              "revoke_record"),
    # ── 失信类（按主体区分）──
    (("个人", "失信"),                           "personal_dishonest_record"),
    (("股东公司", "失信"),                       "shareholder_dishonest_record"),
    (("被投资公司", "失信"),                     "investee_dishonest_record"),
    (("持股", "失信"),                           "investee_dishonest_record"),
    (("失信",),                                  "dishonest_record"),
    # ── 被执行类 ──
    (("个人", "被执行", "详情"),                 "personal_executed_record"),
    (("被执行", "详情"),                         "executed_record"),
    # ── 限制高消费类 ──
    (("个人", "限制高消费", "详情"),             "personal_high_consume_record"),
    (("限制高消费", "详情"),                     "high_consume_record"),
    # ── 财产保全 ──
    (("被保全", "详情"),                         "preservation_record"),
    # ── 终本类 ──
    (("个人", "终本", "详情"),                   "personal_terminated_record"),
    (("终本", "详情"),                           "terminated_record"),
    # ── 司法文书类 ──
    (("裁判文书", "借款合同", "详情"),           "loan_dispute_doc"),
    (("开庭公告", "借款合同", "详情"),           "court_announcement"),
    (("法院公告", "借款合同", "详情"),           "court_notice"),
    # ── 破产/清算 ──
    (("破产", "详情"),                           "bankrupt_record"),
    (("清算", "详情"),                           "bankrupt_record"),
    # ── 税务类 ──
    (("欠税", "详情"),                           "tax_arrears_record"),
    (("税收违法", "详情"),                       "tax_violation_record"),
    # ── 股权类 ──
    (("股权冻结", "详情"),                       "equity_freeze_record"),
    (("股权出质", "详情"),                       "equity_pledge_record"),
    # ── 违法/拍卖 ──
    (("严重违法", "详情"),                       "serious_violation_record"),
    (("司法拍卖", "详情"),                       "judicial_auction_record"),
    # ── 处罚类 ──
    (("行政处罚", "详情"),                       "admin_penalty_record"),
    (("环保处罚", "详情"),                       "env_penalty_record"),
    # ── 经营异常 ──
    (("经营异常", "详情"),                       "biz_abnormal_record"),
    # ── 借贷/担保类 ──
    (("逾期借贷账户", "详情"),                   "overdue_loan_record"),
    (("关注类借贷", "详情"),                     "attention_loan_record"),
    (("逾期对外担保", "详情"),                   "overdue_guarantee_record"),
    (("关注类担保", "详情"),                     "attention_guarantee_record"),
    # ── 案件类 ──
    (("自诉", "详情"),                           "self_prosecution_case"),
    (("为被告", "详情"),                         "defendant_case"),
    # ── 舆情 ──
    (("负面舆情",),                              "negative_news"),
    # ── 资质证书类 ──
    (("资质", "到期"),                           "qualification_near_expiry"),
    (("资质", "天数"),                           "qualification_near_expiry"),
    (("资质", "剩余"),                           "qualification_near_expiry"),
    (("资质", "状态"),                           "qualification_cert"),
    (("资质", "等级"),                           "qualification_cert"),
    # ── 中标 ──
    (("中标",),                                  "bid_record"),
]


# ---------------------------------------------------------------------------
# v2.1.0: 统一模式匹配（消除三处重复）
# ---------------------------------------------------------------------------

def _field_patterns(fname):
    """为指定字段名生成完整的 (正则, 条件类型) 模式列表。
    所有解析函数共用此列表，新增模式只需改一处。"""
    f = re.escape(fname)
    return [
        (rf'{f}\s*>=\s*(\d+)\s*万', 'gte_amount_wan'),
        (rf'{f}\s*≥\s*(\d+)\s*万', 'gte_amount_wan'),
        (rf'{f}\s*>=\s*(\d+(?:\.\d+)?)\s*%', 'gte_percent'),
        (rf'{f}\s*≥\s*(\d+(?:\.\d+)?)\s*%', 'gte_percent'),
        (rf'{f}\s*<\s*(\d+(?:\.\d+)?)\s*%', 'lt_percent'),
        (rf'{f}\s*>\s*(\d+(?:\.\d+)?)\s*万', 'gt_amount_wan'),
        (rf'{f}\s*≤\s*(\d+)\s*天', 'lte_int'),
        (rf'{f}\s*<=\s*(\d+)\s*天', 'lte_int'),
        (rf'{f}\s*≥\s*(\d+)\s*年', 'gte_years'),
        (rf'{f}\s*>=\s*(\d+)\s*年', 'gte_years'),
        (rf'{f}\s*<=\s*(\d+)', 'lte_int'),
        (rf'{f}\s*≤\s*(\d+)', 'lte_int'),
        (rf'{f}\s*<\s*(\d+)', 'lt_int'),
        (rf'{f}\s*>=\s*(\d+)', 'gte_int'),
        (rf'{f}\s*≥\s*(\d+)', 'gte_int'),
        (rf'{f}\s*>\s*(\d+)', 'gt_int'),
        # v2.1.0: negative number support (e.g., = -999)
        (rf'{f}\s*=\s*(-?\d+)', 'eq_int'),
        (rf'{f}\s*<>空', 'not_empty'),
        # v2.1.0: slash-delimited enum (完整捕获所有斜杠分隔值)
        (rf'{f}\s*为/([^/]+(?:/[^/]+)*)', 'slash_enum'),
        # v2.1.0: enum_match 允许"或"和||出现在枚举值内部（如 为d或D）
        (rf'{f}\s*为([^\s且（）]+)', 'enum_match'),
    ]


def _match_patterns(expr, fname, fcode, ftype):
    """尝试用 fname 匹配表达式中的条件，返回 (matched, condition_dict)"""
    for pat, ctype in _field_patterns(fname):
        m = re.search(pat, expr)
        if m:
            val = m.group(1) if m.lastindex else None
            return True, {
                "field_name": fname, "field_code": fcode,
                "field_type": ftype, "operator": ctype,
                "value": val, "condition_type": ctype,
            }
    return False, None


def _normalize_parens(expr):
    """v2.1.0: 将中文全角括号（）归一化为半角()"""
    return expr.replace('（', '(').replace('）', ')')


# ---------------------------------------------------------------------------
# v2.4.2: 阈值/等值比较 fallback 解析器
# 当 parse_expression_conditions 返回 0 条件时兜底，专治 inputParamsMapped
# 不完整（如扁平架构 DF_PRE_CONC_001 的 4 条规则 inputParamsMapped 字段名
# 与 code 映射错位，导致 _match_patterns 全 miss）。
# ---------------------------------------------------------------------------

# 比较运算符 → condition_type 前缀映射
_CMP_OPS = [
    (">=", "gte"), ("<=", "lte"), (">", "gt"), ("<", "lt"),
    ("<>", "neq"), ("=", "eq"),
]


def _extract_threshold_fallback(expression, fields_dict):
    """正则兜底提取阈值/等值条件。

    Args:
        expression: 规则表达式（原始文本）
        fields_dict: 字段字典 {field_name: {code, type, ...}}，用于按名查 code

    Returns:
        list[dict]：结构化条件列表，结构与 parse_expression_conditions 一致。
    """
    if not fields_dict:
        return []
    conditions = []
    seen = set()
    expr = _normalize_parens(expression)
    # 移除"条件1："、"1："等前缀，避免干扰
    expr_clean = re.sub(r'(?:条件\s*\d+\s*[:：])|(?:\d+\s*[:：])', '', expr)

    # v2.4.2: 前缀模糊匹配——表达式中的截断字段名 → fields_dict 中的完整字段名
    _field_prefix_cache = {}
    def _resolve_field(name):
        """精确匹配优先，然后前缀匹配（取最短匹配避免歧义）。"""
        if name in fields_dict:
            return name
        if name in _field_prefix_cache:
            return _field_prefix_cache[name]
        candidates = [k for k in fields_dict if k.startswith(name) and len(k) > len(name)]
        result = min(candidates, key=len) if candidates else None
        _field_prefix_cache[name] = result
        return result

    # 1) 字段-字段比较：X >= Y / X <= Y / X > Y / X < Y / X = Y / X <> Y
    #    v2.4.2: 右值必须也在 fields_dict 中才算字段-字段比较，否则留给字符串等值分支
    for op, prefix in _CMP_OPS:
        op_esc = re.escape(op)
        pat = rf'([\w\u4e00-\u9fa5_]+)\s*{op_esc}\s*([\w\u4e00-\u9fa5_]+)'
        for m in re.finditer(pat, expr_clean):
            left_raw, right_raw = m.group(1).strip(), m.group(2).strip()
            # 过滤纯数字右值（那是常量比较，走下面分支）
            if re.fullmatch(r'\d+(\.\d+)?', right_raw):
                continue
            # v2.4.2: 用模糊匹配解析字段名
            left = _resolve_field(left_raw)
            right = _resolve_field(right_raw)
            # 右值必须是已知字段，否则让给字符串等值分支
            if right is None:
                continue
            if left is None or (left, right) in seen:
                continue
            seen.add((left, right))
            left_meta = fields_dict[left]
            right_meta = fields_dict.get(right, {})
            conditions.append({
                "field_name": left,
                "field_code": left_meta.get("code"),
                "field_type": left_meta.get("type", "小数型"),
                "operator": op,
                "value": right,
                "value_field_code": right_meta.get("code"),
                "value_field_type": right_meta.get("type"),
                "condition_type": f"{prefix}_field",
            })

    # 2) 字段-常量比较：X >= 100 / X = "重庆" 等
    #    v2.4.2: 用模糊匹配解析字段名（表达式可能用截断名）
    for op, prefix in _CMP_OPS:
        op_esc = re.escape(op)
        pat = rf'([\w\u4e00-\u9fa5_]+)\s*{op_esc}\s*(\d+(?:\.\d+)?)'
        for m in re.finditer(pat, expr_clean):
            fname_raw, val = m.group(1).strip(), m.group(2)
            fname = _resolve_field(fname_raw)
            if fname is None or (fname, val) in seen:
                continue
            seen.add((fname, val))
            meta = fields_dict[fname]
            conditions.append({
                "field_name": fname,
                "field_code": meta.get("code"),
                "field_type": meta.get("type", "小数型"),
                "operator": op,
                "value": val,
                "condition_type": f"{prefix}_int" if "." not in val else f"{prefix}_float",
            })

    # 3) 字符串等值：字段 = "重庆" / 字段 <> "成都"（引号/非引号都兼容）
    #    v2.4.2: 用模糊匹配解析字段名
    str_pat = re.compile(
        r'([\w\u4e00-\u9fa5_]+)\s*(=|<>)\s*["\']?([\u4e00-\u9fa5A-Za-z][\w\u4e00-\u9fa5]*)["\']?'
    )
    for m in str_pat.finditer(expr_clean):
        fname_raw, op, val = m.group(1).strip(), m.group(2), m.group(3).strip()
        # 跳过数值（已被上面分支处理）
        if re.fullmatch(r'\d+(\.\d+)?', val):
            continue
        fname = _resolve_field(fname_raw)
        if fname is None or (fname, val) in seen:
            continue
        seen.add((fname, val))
        meta = fields_dict[fname]
        prefix = "eq" if op == "=" else "neq"
        conditions.append({
            "field_name": fname,
            "field_code": meta.get("code"),
            "field_type": meta.get("type", "字符型"),
            "operator": op,
            "value": val,
            "condition_type": f"{prefix}_str",
        })

    return [c for c in conditions if c.get("field_code")]


# ---------------------------------------------------------------------------
# 表达式解析与参数值构造
# ---------------------------------------------------------------------------

def parse_expression_conditions(expression, input_params_mapped):
    """解析规则表达式，提取条件列表。"""
    conditions = []
    expr = _normalize_parens(expression.strip())

    expanded_params = []
    for p in input_params_mapped:
        name = p["name"]
        if "\n" in name:
            for sub_name in name.split("\n"):
                sub_name = sub_name.strip()
                if sub_name:
                    expanded_params.append({
                        "name": sub_name, "code": p.get("code"),
                        "type": p.get("type", "整数型"), "fuzzyMatch": sub_name,
                    })
        else:
            expanded_params.append(p)

    for p in expanded_params:
        fname = p["name"]
        fcode = p.get("code")
        ftype = p.get("type", "整数型")
        if not fcode:
            continue

        matched, cond = _match_patterns(expr, fname, fcode, ftype)
        if matched:
            conditions.append(cond)
            continue

        if "_" in fname:
            fname_norm = fname.replace("_", "")
            matched, cond = _match_patterns(expr, fname_norm, fcode, ftype)
            if matched:
                cond["field_name"] = fname
                conditions.append(cond)

    return conditions


def construct_param_value(condition, mode):
    """根据条件和模式(hit/miss/boundary)构造参数值。"""
    ctype = condition["condition_type"]
    fname = condition["field_name"]
    ftype = condition["field_type"]
    val = condition.get("value")

    if ctype == "gt_int" and val == "0":
        return 3 if mode == "hit" else (0 if mode == "miss" else 1)

    elif ctype == "not_empty":
        if mode == "miss":
            return "[]"
        return get_detail_template(fname)

    elif ctype == "gte_percent":
        threshold = float(val)
        if mode == "hit": return round(threshold + 10, 1)
        elif mode == "miss": return round(max(0, threshold - 15), 1)
        return threshold

    elif ctype == "lt_percent":
        threshold = float(val)
        if mode == "hit": return round(max(0, threshold - 5), 1)
        elif mode == "miss": return round(threshold + 10, 1)
        return round(threshold - 0.1, 2)

    elif ctype == "gte_amount_wan":
        t = int(val)
        if mode == "hit": return t + 100
        elif mode == "miss": return max(0, t - 100)
        return t

    elif ctype == "gt_amount_wan":
        t = int(val)
        if mode == "hit": return t + 100
        elif mode == "miss": return t
        return t + 1

    elif ctype == "gte_int":
        t = int(val)
        if t == 0:
            return 3 if mode == "hit" else (0 if mode == "miss" else 1)
        if mode == "hit": return t + 1
        elif mode == "miss": return t - 1
        return t

    elif ctype == "gt_int":
        t = int(val)
        if mode == "hit": return t + 3
        elif mode == "miss": return t
        return t + 1

    elif ctype == "lte_int":
        t = int(val)
        if mode == "hit": return t - 1
        elif mode == "miss": return t + 5
        return t

    elif ctype == "lt_int":
        t = int(val)
        if mode == "hit": return t - 1
        elif mode == "miss": return t + 1
        return t - 1

    elif ctype == "eq_int":
        t = int(val)
        if mode == "hit": return t
        elif mode == "miss": return t + 5
        return t

    elif ctype == "enum_match":
        all_vals = _split_enum_values(val)
        if mode == "hit": return all_vals[0]
        elif mode == "miss": return "其他值"
        return all_vals[-1] if len(all_vals) > 1 else all_vals[0]

    elif ctype == "slash_enum":
        all_vals = _parse_slash_enum_values(val)
        if mode == "hit": return all_vals[0]
        elif mode == "miss": return "正常"
        return all_vals[-1] if len(all_vals) > 1 else all_vals[0]

    elif ctype == "gte_years":
        years = int(val)
        today = datetime.now()
        if mode == "hit":
            return _safe_date_years_ago(today, years + 1).strftime("%Y-%m-%d")
        elif mode == "miss":
            return _safe_date_years_ago(today, max(years - 1, 0)).strftime("%Y-%m-%d")
        return _safe_date_years_ago(today, years).strftime("%Y-%m-%d")

    # v2.4.2: fallback 解析器产出的 condition_type（字段-字段 / 字段-字符串 / 字段-浮点常量）
    # field-to-field 比较：只能控制 LHS 字段值，RHS 由测试环境提供。启发式 500万/50万/100万。
    if ctype == "gte_field":
        return 5000000 if mode == "hit" else (500000 if mode == "miss" else 1000000)
    if ctype == "gt_field":
        return 5000000 if mode == "hit" else (500000 if mode == "miss" else 1000001)
    if ctype == "lte_field":
        return 500000 if mode == "hit" else (5000000 if mode == "miss" else 1000000)
    if ctype == "lt_field":
        return 500000 if mode == "hit" else (5000000 if mode == "miss" else 999999)
    if ctype == "eq_field":
        return 1000000 if mode in ("hit", "boundary") else 500000
    if ctype == "neq_field":
        # 字符串不等：hit 给一个不在常见枚举中的值，miss 给一个典型值
        if str(val) in ("重庆", "四川", "成都") or any("\u4e00" <= ch <= "\u9fa5" for ch in str(val)):
            return "其他省" if mode == "hit" else val
        return 999 if mode == "hit" else 1000000
    if ctype == "eq_str":
        return val if mode in ("hit", "boundary") else f"非{val}"
    if ctype == "neq_str":
        return f"非{val}" if mode == "hit" else val
    if ctype in ("gte_float", "gt_float"):
        t = float(val)
        if mode == "hit": return round(t + max(t * 0.2, 1), 2)
        if mode == "miss": return round(max(0, t * 0.5), 2)
        return t
    if ctype in ("lte_float", "lt_float"):
        t = float(val)
        if mode == "hit": return round(max(0, t * 0.5), 2)
        if mode == "miss": return round(t + max(t * 0.2, 1), 2)
        return t
    if ctype == "eq_int":  # fallback 路径的 eq_int（常量比较）
        t = int(val)
        return t if mode in ("hit", "boundary") else t + 5

    # fallback by type
    if ftype == "整数型":
        return 1 if mode == "hit" else 0
    elif ftype == "小数型":
        return 0.5 if mode == "hit" else 0.0
    return "测试值" if mode == "hit" else ""


def _split_enum_values(enum_val):
    """v2.1.0: 将枚举值字符串按 或/|| 拆分为列表。"""
    if not enum_val:
        return [enum_val]
    parts = re.split(r'\|\||或', enum_val)
    return [p.strip() for p in parts if p.strip()] or [enum_val]


def _parse_slash_enum_values(raw):
    """v2.1.0: 解析斜杠分隔的枚举值。如 '吊销/或/撤销' → ['吊销', '撤销']"""
    if not raw:
        return [raw]
    parts = [p.strip() for p in raw.split('/') if p.strip() and p.strip() != '或']
    return parts if parts else [raw]


def _safe_date_years_ago(base_date, years):
    """v2.1.0: 安全地计算 N 年前的日期，处理闰年 2/29 问题。"""
    target_year = base_date.year - years
    try:
        return base_date.replace(year=target_year)
    except ValueError:
        return base_date.replace(year=target_year, day=28)


def _safe_date_days_ago(base_date, days):
    """v2.2.0: 安全地计算 N 天前的日期。"""
    return base_date - timedelta(days=days)


def _extract_day_window(expression):
    """v2.2.0: 从表达式中提取 '距今≤N天' 的天数阈值。
    匹配模式：≤N天、<=N天、<N天。返回 (天数, 运算符) 或 None。"""
    m = re.search(r'[≤<]=?\s*(\d+)\s*天', expression)
    if m:
        days = int(m.group(1))
        op = "lt" if "≤" not in m.group(0) and "<" in m.group(0) and "=" not in m.group(0) else "lte"
        return days, op
    return None


def generate_time_boundary_cases(code, name, rs_code, hit_result, expression, conditions):
    """v2.2.0: 为含时间窗口条件的规则生成日期边界用例。
    专门测试'距今≤N天'类条件的日期临界值。"""
    day_window = _extract_day_window(expression)
    if not day_window:
        return []

    days, op = day_window
    today = datetime.now()
    cases = []

    # 确定详情字段 — 找到 not_empty 条件对应的字段
    detail_cond = None
    count_cond = None
    for c in conditions:
        if c["condition_type"] == "not_empty":
            detail_cond = c
        elif c["condition_type"] in ("gt_int", "gte_int") and c.get("value") == "0":
            count_cond = c

    if not detail_cond:
        return []

    # ── 边界命中：记录恰好在窗口内（N-1天前 / N天前）──
    params_hit = {}
    for k, v in COMMON_PARAMS.items():
        params_hit[k] = v.format(tc_id=f"{code}_TBH")
    # 设置 count > 0
    if count_cond:
        params_hit[count_cond["field_code"]] = 1
    # 设置详情，日期恰好=窗口边界
    boundary_date = _safe_date_days_ago(today, days).strftime("%Y-%m-%d")
    detail_tpl = json.loads(get_detail_template(detail_cond["field_name"]))
    if detail_tpl and isinstance(detail_tpl, list) and len(detail_tpl) > 0:
        # 找到日期字段并设置为边界日期
        for key in detail_tpl[0]:
            if "date" in key.lower() or "Date" in key or "日期" in key:
                detail_tpl[0][key] = boundary_date
                break
    params_hit[detail_cond["field_code"]] = json.dumps(detail_tpl, ensure_ascii=False)

    op_desc = "≤" if op == "lte" else "<"
    cases.append({
        "id": f"TC_{code}_TBH",
        "desc": f"时间窗口边界命中：{name} - 记录恰好{days}天前",
        "group": "规则", "caseType": "边界案例",
        "scenario": f"时间窗口边界：记录日期={boundary_date}(恰好{days}天前)，{op_desc}{days}天条件应命中",
        "expected": f"命中：{code}({name})[{hit_result}]",
        "params": params_hit, "targetRuleSet": rs_code, "targetRule": code,
    })

    # ── 边界未命中：记录刚好在窗口外（N+1天前）──
    params_miss = {}
    for k, v in COMMON_PARAMS.items():
        params_miss[k] = v.format(tc_id=f"{code}_TBM")
    if count_cond:
        params_miss[count_cond["field_code"]] = 1
    outside_date = _safe_date_days_ago(today, days + 1).strftime("%Y-%m-%d")
    detail_tpl_miss = json.loads(get_detail_template(detail_cond["field_name"]))
    if detail_tpl_miss and isinstance(detail_tpl_miss, list) and len(detail_tpl_miss) > 0:
        for key in detail_tpl_miss[0]:
            if "date" in key.lower() or "Date" in key or "日期" in key:
                detail_tpl_miss[0][key] = outside_date
                break
    params_miss[detail_cond["field_code"]] = json.dumps(detail_tpl_miss, ensure_ascii=False)

    cases.append({
        "id": f"TC_{code}_TBM",
        "desc": f"时间窗口边界未命中：{name} - 记录恰好{days+1}天前",
        "group": "规则", "caseType": "边界案例",
        "scenario": f"时间窗口边界：记录日期={outside_date}(恰好{days+1}天前)，超出{op_desc}{days}天窗口应不命中",
        "expected": f"未命中：{code}({name})",
        "params": params_miss, "targetRuleSet": rs_code, "targetRule": code,
    })

    return cases


def has_threshold_condition(conditions):
    """检查是否有阈值条件（需要生成边界用例）"""
    threshold_types = {"gte_percent", "lt_percent", "gte_amount_wan", "gt_amount_wan",
                       "gte_int", "gt_int", "lte_int", "lt_int", "eq_int", "gte_years",
                       # v2.4.2: fallback 解析器产出的数值型 condition_type
                       "gte_field", "lte_field", "gte_float", "lte_float"}
    return any(c["condition_type"] in threshold_types and c.get("value") not in ("0", None) for c in conditions)


# ---------------------------------------------------------------------------
# AND/OR 条件拆解
# ---------------------------------------------------------------------------

def is_strict_operator(condition_type):
    return condition_type in ("gt_int", "lt_int", "gt_amount_wan",
                              "gt_field", "lt_field", "gt_float", "lt_float")


def has_or_connection(expression):
    """v2.1.0: 检测表达式是否包含 OR 条件（字段间的"或"或"||"）
    排除枚举值内的"或"（如"为d或D"）。"""
    if not expression:
        return False
    expr = _normalize_parens(expression)
    # 移除枚举值段
    cleaned = re.sub(r'为(?:[^\s且|（）]+(?:或|\|\|))*[^\s且|（）]+', '', expr)
    if "或" in cleaned:
        return True
    if "||" in cleaned:
        return True
    return False


def split_or_branches(expression, input_params_mapped):
    """v2.1.0: 将 OR 表达式拆分为独立分支。
    支持"或"和"||"，正确处理括号嵌套。"""
    expr = _normalize_parens(expression.strip())

    expanded_params = []
    for p in input_params_mapped:
        name = p["name"]
        if "\n" in name:
            for sub_name in name.split("\n"):
                sub_name = sub_name.strip()
                if sub_name:
                    expanded_params.append({"name": sub_name, "code": p.get("code"), "type": p.get("type", "整数型")})
        else:
            expanded_params.append(p)

    field_positions = []
    for p in expanded_params:
        fname = p["name"]
        fcode = p.get("code")
        if not fcode:
            continue
        pos = expr.find(fname)
        if pos >= 0:
            field_positions.append((pos, p))

    if len(field_positions) < 2:
        return [parse_expression_conditions(expression, input_params_mapped)]

    field_positions.sort(key=lambda x: x[0])

    branches = []
    current_branch_fields = [field_positions[0]]

    for i in range(1, len(field_positions)):
        prev_end = current_branch_fields[-1][0] + len(current_branch_fields[-1][1]["name"])
        curr_start = field_positions[i][0]
        between = expr[prev_end:curr_start]
        between_clean = re.sub(r'为(?:[^\s且|（）]+(?:或|\|\|))*[^\s且|（）]+', '', between)

        # 检查括号深度：只在最外层 OR 处拆分
        depth = 0
        has_or_at_top = False
        j = 0
        while j < len(between_clean):
            ch = between_clean[j]
            if ch == '(':
                depth += 1
            elif ch == ')':
                depth -= 1
            elif depth == 0:
                if ch == '或':
                    has_or_at_top = True
                    break
                if ch == '|' and j + 1 < len(between_clean) and between_clean[j + 1] == '|':
                    has_or_at_top = True
                    break
            j += 1

        if has_or_at_top:
            branches.append(current_branch_fields)
            current_branch_fields = [field_positions[i]]
        else:
            current_branch_fields.append(field_positions[i])

    branches.append(current_branch_fields)

    result = []
    for branch_fields in branches:
        branch_conditions = []
        for _, p in branch_fields:
            fname = p["name"]
            fcode = p.get("code")
            ftype = p.get("type", "整数型")
            matched, cond = _match_patterns(expr, fname, fcode, ftype)
            if matched:
                branch_conditions.append(cond)
            elif "_" in fname:
                matched, cond = _match_patterns(expr, fname.replace("_", ""), fcode, ftype)
                if matched:
                    cond["field_name"] = fname
                    branch_conditions.append(cond)
        if branch_conditions:
            result.append(branch_conditions)

    return result if result else [parse_expression_conditions(expression, input_params_mapped)]


def generate_and_decomposition_cases(code, name, rs_code, hit_result, conditions):
    cases = []
    if len(conditions) < 2:
        return cases

    for i, target_cond in enumerate(conditions):
        if target_cond["condition_type"] in ("not_empty", "enum_match"):
            continue

        params = {}
        for k, v in COMMON_PARAMS.items():
            params[k] = v.format(tc_id=f"{code}_AND{i}")

        for j, cond in enumerate(conditions):
            mode = "miss" if j == i else "hit"
            params[cond["field_code"]] = construct_param_value(cond, mode)

        cases.append({
            "id": f"TC_{code}_AND{i}",
            "desc": f"AND拆解：{name} - 仅{target_cond['field_name']}不满足",
            "group": "规则", "caseType": "反案例",
            "scenario": f"AND条件拆解：仅第{i+1}个条件({target_cond['field_name']})不满足，其余满足，验证不命中",
            "expected": f"未命中：{code}({name})",
            "params": params, "targetRuleSet": rs_code, "targetRule": code,
        })

    return cases


def generate_or_branch_cases(code, name, rs_code, hit_result, or_branches):
    cases = []
    if len(or_branches) < 2:
        return cases

    for i, branch in enumerate(or_branches):
        if not branch:
            continue

        params_hit = {}
        for k, v in COMMON_PARAMS.items():
            params_hit[k] = v.format(tc_id=f"{code}_OR{i}_H")

        for j, other_branch in enumerate(or_branches):
            mode = "hit" if j == i else "miss"
            for cond in other_branch:
                params_hit[cond["field_code"]] = construct_param_value(cond, mode)

        branch_desc = "、".join(c["field_name"] for c in branch)
        cases.append({
            "id": f"TC_{code}_OR{i}_HIT",
            "desc": f"OR分支命中：{name} - 仅满足{branch_desc}",
            "group": "规则", "caseType": "正案例",
            "scenario": f"OR条件分支：仅第{i+1}个分支满足，验证规则命中 {hit_result}",
            "expected": f"命中：{code}({name})[{hit_result}]",
            "params": params_hit, "targetRuleSet": rs_code, "targetRule": code,
        })

        has_boundary = any(c["condition_type"] in ("gte_int", "gte_percent", "gte_amount_wan", "gt_int", "lt_int", "lt_percent") for c in branch)
        if has_boundary:
            params_bound = {}
            for k, v in COMMON_PARAMS.items():
                params_bound[k] = v.format(tc_id=f"{code}_OR{i}_B")
            for j, other_branch in enumerate(or_branches):
                mode = "boundary" if j == i else "miss"
                for cond in other_branch:
                    params_bound[cond["field_code"]] = construct_param_value(cond, mode)
            cases.append({
                "id": f"TC_{code}_OR{i}_BOUND",
                "desc": f"OR分支边界：{name} - {branch_desc}在阈值上",
                "group": "规则", "caseType": "边界案例",
                "scenario": f"OR条件边界：第{i+1}个分支在精确阈值上，验证包含/不包含等于",
                "expected": f"命中：{code}({name})[{hit_result}]",
                "params": params_bound, "targetRuleSet": rs_code, "targetRule": code,
            })

    params_all_miss = {}
    for k, v in COMMON_PARAMS.items():
        params_all_miss[k] = v.format(tc_id=f"{code}_OR_MISS")
    for branch in or_branches:
        for cond in branch:
            params_all_miss[cond["field_code"]] = construct_param_value(cond, "miss")

    cases.append({
        "id": f"TC_{code}_OR_MISS",
        "desc": f"OR全不满足：{name} - 所有分支均不满足",
        "group": "规则", "caseType": "反案例",
        "scenario": "OR条件全不满足：所有分支均不满足，验证不命中",
        "expected": f"未命中：{code}({name})",
        "params": params_all_miss, "targetRuleSet": rs_code, "targetRule": code,
    })

    return cases


def generate_strict_boundary_cases(code, name, rs_code, hit_result, conditions):
    cases = []
    for i, cond in enumerate(conditions):
        if not is_strict_operator(cond["condition_type"]):
            continue
        val = cond.get("value")
        if val is None:
            continue

        params = {}
        for k, v in COMMON_PARAMS.items():
            params[k] = v.format(tc_id=f"{code}_STRICT{i}")

        for j, other_cond in enumerate(conditions):
            if j == i:
                params[cond["field_code"]] = int(val)
            else:
                params[other_cond["field_code"]] = construct_param_value(other_cond, "hit")

        op_symbol = ">" if "gt" in cond["condition_type"] else "<"
        cases.append({
            "id": f"TC_{code}_STRICT{i}",
            "desc": f"严格边界：{name} - {cond['field_name']}={val}（{op_symbol}不含等于）",
            "group": "规则", "caseType": "边界案例",
            "scenario": f"严格运算符边界：{cond['field_name']}恰好={val}，{op_symbol}不含等于，应不命中",
            "expected": f"未命中：{code}({name})",
            "params": params, "targetRuleSet": rs_code, "targetRule": code,
        })

    return cases


# ---------------------------------------------------------------------------
# 规则级用例生成
# ---------------------------------------------------------------------------

def generate_rule_test_cases(rules, fields):
    test_cases = []

    for rule in rules:
        code = rule["code"]
        name = rule["name"]
        rs_code = rule["ruleSetCode"]
        hit_result = rule["hitResult"]
        expression = rule["expression"]

        if rs_code == "alertLevelJudgment":
            continue

        conditions = parse_expression_conditions(expression, rule.get("inputParamsMapped", []))

        # v2.4.2: 主解析器返回 0 条件时调用正则 fallback（修扁平架构 inputParamsMapped 不完整问题）
        if not conditions and isinstance(fields, dict):
            fb_conds = _extract_threshold_fallback(expression, fields)
            if fb_conds:
                print(f"[{code}] fallback 解析出 {len(fb_conds)} 个条件（主解析器 0）", file=sys.stderr)
                conditions = fb_conds

        # 正案例
        params = {}
        for k, v in COMMON_PARAMS.items():
            params[k] = v.format(tc_id=code)
        for cond in conditions:
            params[cond["field_code"]] = construct_param_value(cond, "hit")

        test_cases.append({
            "id": f"TC_{code}_HIT", "desc": f"命中验证：{name}",
            "group": "规则", "caseType": "正案例",
            "scenario": f"构造满足 {code} 全部条件的参数值，验证规则命中 {hit_result}",
            "expected": f"命中：{code}({name})[{hit_result}]",
            "params": params, "targetRuleSet": rs_code, "targetRule": code,
        })

        # 反案例
        params_miss = {}
        for k, v in COMMON_PARAMS.items():
            params_miss[k] = v.format(tc_id=f"{code}_M")
        for cond in conditions:
            params_miss[cond["field_code"]] = construct_param_value(cond, "miss")

        test_cases.append({
            "id": f"TC_{code}_MISS", "desc": f"未命中验证：{name}",
            "group": "规则", "caseType": "反案例",
            "scenario": f"构造不满足 {code} 条件的参数值，验证规则不命中",
            "expected": f"未命中：{code}({name})",
            "params": params_miss, "targetRuleSet": rs_code, "targetRule": code,
        })

        # 边界值
        if has_threshold_condition(conditions):
            params_bound = {}
            for k, v in COMMON_PARAMS.items():
                params_bound[k] = v.format(tc_id=f"{code}_B")
            for cond in conditions:
                params_bound[cond["field_code"]] = construct_param_value(cond, "boundary")

            test_cases.append({
                "id": f"TC_{code}_BOUND", "desc": f"边界值验证：{name}",
                "group": "规则", "caseType": "边界案例",
                "scenario": f"使用阈值临界值测试 {code}，验证边界判定",
                "expected": f"命中：{code}({name})[{hit_result}]",
                "params": params_bound, "targetRuleSet": rs_code, "targetRule": code,
            })

        # AND 拆解（v2.4.2: 跳过有 field_code 碰撞的条件组，避免 params 被覆盖）
        if len(conditions) >= 2 and not has_or_connection(expression):
            field_codes = [c["field_code"] for c in conditions]
            has_collision = len(field_codes) != len(set(field_codes))
            if not has_collision:
                test_cases.extend(generate_and_decomposition_cases(code, name, rs_code, hit_result, conditions))
            else:
                print(f"[{code}] AND拆解跳过：{len(field_codes)-len(set(field_codes))} 组 field_code 碰撞", file=sys.stderr)

        # OR 拆解
        if has_or_connection(expression):
            or_branches = split_or_branches(expression, rule.get("inputParamsMapped", []))
            if len(or_branches) >= 2:
                test_cases.extend(generate_or_branch_cases(code, name, rs_code, hit_result, or_branches))

        # 严格边界
        test_cases.extend(generate_strict_boundary_cases(code, name, rs_code, hit_result, conditions))

        # v2.2.0: 时间窗口边界
        test_cases.extend(generate_time_boundary_cases(code, name, rs_code, hit_result, expression, conditions))

    return test_cases


# ---------------------------------------------------------------------------
# 函数级用例生成（v2.1.0: 模糊名称匹配回退）
# ---------------------------------------------------------------------------

FUNC_CASE_MAP = {
    "初始化函数": [
        {"caseType": "正案例", "scenario": "传入有效 AppKey 和 SecretKey，验证企查查接口可正常调用", "expected": "函数输出：初始化函数+企查查appKey=有效值，企查查SecretKey=有效值"},
        {"caseType": "反案例", "scenario": "传入空 AppKey，验证初始化异常处理", "expected": "函数输出：初始化函数+初始化失败"},
    ],
    "预警等级计算": [
        {"caseType": "正案例", "scenario": "高风险命中条数=1，中风险命中条数=2，低风险命中条数=3", "expected": "函数输出：预警等级计算+高风险规则命中条数=1，中风险规则命中条数=2，低风险规则命中条数=3"},
        {"caseType": "正案例", "scenario": "所有等级命中条数均为0", "expected": "函数输出：预警等级计算+极高风险规则命中条数=0，高风险规则命中条数=0，中风险规则命中条数=0，低风险规则命中条数=0"},
        {"caseType": "边界案例", "scenario": "中风险命中条数=3（触发红色预警阈值）", "expected": "函数输出：预警等级计算+中风险规则命中条数=3"},
        {"caseType": "边界案例", "scenario": "中风险命中条数=2（未达红色预警阈值，应为黄色预警）", "expected": "函数输出：预警等级计算+中风险规则命中条数=2"},
        {"caseType": "正案例", "scenario": "HIGH=4, MED=2, LOW=0, EXTREME=0 混合计数验证", "expected": "函数输出：预警等级计算+极高风险规则命中条数=0，高风险规则命中条数=4，中风险规则命中条数=2，低风险规则命中条数=0"},
        {"caseType": "正案例", "scenario": "HIGH=9, MED=7, LOW=0, EXTREME=0 大量混合计数验证", "expected": "函数输出：预警等级计算+极高风险规则命中条数=0，高风险规则命中条数=9，中风险规则命中条数=7，低风险规则命中条数=0"},
        {"caseType": "边界案例", "scenario": "空字符串输入(所有四级计数=0)", "expected": "函数输出：预警等级计算+极高风险规则命中条数=0，高风险规则命中条数=0，中风险规则命中条数=0，低风险规则命中条数=0"},
        {"caseType": "正案例", "scenario": "极高风险>0(EBNP08资质已过期触发极高风险预警)", "expected": "函数输出：预警等级计算+极高风险规则命中条数=1"},
    ],
    "股东公司循环函数": [
        {"caseType": "正案例", "scenario": "持股超20%股东公司3家，循环下标从0开始递增", "expected": "函数输出：股东公司循环函数+股东公司名称=第N家公司名，股东公司循环函数结束标识=否"},
        {"caseType": "边界案例", "scenario": "循环下标=公司列表长度，验证循环终止", "expected": "函数输出：股东公司循环函数+股东公司循环函数结束标识=是"},
        {"caseType": "反案例", "scenario": "持股超20%股东公司列表为空", "expected": "函数输出：股东公司循环函数+股东公司循环函数结束标识=是"},
        {"caseType": "正案例", "scenario": "多公司(JSON数组格式)", "expected": "函数输出：股东公司循环函数+主要股东公司keyno列表=[COMP_A,COMP_B,COMP_C]，循环索引=3"},
        {"caseType": "正案例", "scenario": "多公司(英文逗号分隔)", "expected": "函数输出：股东公司循环函数+主要股东公司keyno列表=COMP_A,COMP_B,COMP_C，循环索引=3"},
        {"caseType": "正案例", "scenario": "多公司(中文逗号分隔)", "expected": "函数输出：股东公司循环函数+主要股东公司keyno列表=COMP_A，COMP_B，COMP_C，循环索引=3"},
        {"caseType": "正案例", "scenario": "单公司(1个keyno)单次迭代后结束", "expected": "函数输出：股东公司循环函数+主要股东公司keyno列表=COMP_A，循环索引=1，结束标识=是"},
    ],
    "持股公司循环函数": [
        {"caseType": "正案例", "scenario": "对外投资持股超20%公司2家，循环下标递增", "expected": "函数输出：持股公司循环函数+持股公司名称=第N家公司名，持股公司循环函数结束标识=否"},
        {"caseType": "边界案例", "scenario": "循环下标=公司列表长度，验证循环终止", "expected": "函数输出：持股公司循环函数+持股公司循环函数结束标识=是"},
        {"caseType": "正案例", "scenario": "多公司(JSON数组格式)", "expected": "函数输出：持股公司循环函数+持股公司keyno列表=[COMP_A,COMP_B,COMP_C]，循环索引=3"},
        {"caseType": "正案例", "scenario": "多公司(英文逗号分隔)", "expected": "函数输出：持股公司循环函数+持股公司keyno列表=COMP_A,COMP_B,COMP_C，循环索引=3"},
        {"caseType": "正案例", "scenario": "多公司(中文逗号分隔)", "expected": "函数输出：持股公司循环函数+持股公司keyno列表=COMP_A，COMP_B，COMP_C，循环索引=3"},
        {"caseType": "正案例", "scenario": "单公司(1个keyno)单次迭代后结束", "expected": "函数输出：持股公司循环函数+持股公司keyno列表=[COMP_A]，循环索引=1，结束标识=是"},
        {"caseType": "正案例", "scenario": "跨迭代失信计数累加(2+3=5)和详情合并", "expected": "函数输出：持股公司循环函数+被投资公司失信记录累加=5，详情数组合并"},
    ],
    "股东公司数据合并": [
        {"caseType": "正案例", "scenario": "3家股东公司各有失信记录，验证数据合并结果", "expected": "函数输出：股东公司数据合并+合并后失信记录总数=三家之和"},
        {"caseType": "反案例", "scenario": "所有股东公司均无失信记录", "expected": "函数输出：股东公司数据合并+合并后失信记录总数=0"},
        {"caseType": "正案例", "scenario": "首次迭代初始累加(0+2=2)和详情写入", "expected": "函数输出：股东公司数据合并+失信标记=2(循环迭代2次)"},
        {"caseType": "正案例", "scenario": "跨迭代累加(2+3=5)和详情数组合并", "expected": "函数输出：股东公司数据合并+失信标记=5(循环迭代5次)"},
        {"caseType": "边界案例", "scenario": "临时字段为null/空(安全处理不崩溃)", "expected": "函数输出：股东公司数据合并+失信标记=0(null/空安全处理)"},
    ],
    "持股公司数据合并": [
        {"caseType": "正案例", "scenario": "2家持股公司各有失信记录，验证数据合并", "expected": "函数输出：持股公司数据合并+合并后失信记录总数=两家之和"},
        {"caseType": "反案例", "scenario": "持股公司列表为空", "expected": "函数输出：持股公司数据合并+合并后失信记录总数=0"},
        {"caseType": "正案例", "scenario": "首次迭代初始累加(0+1=1)和详情写入", "expected": "函数输出：持股公司数据合并+失信标记=1(循环迭代1次)"},
        {"caseType": "正案例", "scenario": "跨迭代累加(2+3=5)和详情数组合并", "expected": "函数输出：持股公司数据合并+失信标记=5(循环迭代5次)"},
        {"caseType": "边界案例", "scenario": "临时字段为null/空(安全处理不崩溃)", "expected": "函数输出：持股公司数据合并+失信标记=0(null/空安全处理)"},
    ],
    "企业信息对象取值": [
        {"caseType": "正案例", "scenario": "传入有效企业信息对象", "expected": "函数输出：企业信息对象取值+企业名称=测试企业，统一社会信用代码=91XXXXXXXX"},
        {"caseType": "反案例", "scenario": "传入空企业信息对象", "expected": "函数输出：企业信息对象取值+企业名称=空"},
        {"caseType": "边界案例", "scenario": "C_O_ENTERPRISEINFO空数组[]输入", "expected": "函数返回空(空数组输入)，企业子策略跳过"},
        {"caseType": "边界案例", "scenario": "C_O_ENTERPRISEINFO=null输入", "expected": "企业子策略跳过(null输入安全处理)"},
        {"caseType": "正案例", "scenario": "多对象数组(仅取第一个元素)", "expected": "函数输出：企业信息对象取值+企业名称=第一企业(仅取第一个元素)"},
    ],
    "个人信息对象取值": [
        {"caseType": "正案例", "scenario": "传入有效个人信息对象", "expected": "函数输出：个人信息对象取值+企业人员姓名=张三，个人角色类型=法定代表人"},
        {"caseType": "反案例", "scenario": "传入空个人信息对象", "expected": "函数输出：个人信息对象取值+企业人员姓名=空"},
        {"caseType": "边界案例", "scenario": "C_O_PERSONINFO空数组输入不崩溃", "expected": "函数返回空; 个人子策略跳过(不崩溃)"},
        {"caseType": "正案例", "scenario": "多对象数组(仅取第一个元素)", "expected": "函数输出：个人信息对象取值+企业人员姓名=张三(仅取第一个元素)"},
    ],
    "关联企业信息对象取值": [
        {"caseType": "正案例", "scenario": "传入有效关联企业信息对象", "expected": "函数输出：关联企业信息对象取值+企业名称=关联测试企业"},
        {"caseType": "反案例", "scenario": "传入空关联企业信息对象", "expected": "函数输出：关联企业信息对象取值+企业名称=空"},
        {"caseType": "边界案例", "scenario": "C_O_RELATEDCOMPANYINFO空数组输入不崩溃", "expected": "关联企业子策略跳过; 策略整体不受影响"},
        {"caseType": "正案例", "scenario": "多对象数组(仅取第一个元素)", "expected": "函数输出：关联企业信息对象取值+企业名称=关联A(仅取第一个元素)"},
    ],
    "资质等级计算函数": [
        {"caseType": "正案例", "scenario": "准入基线=特级，当前资质=一级，验证降级档位=1", "expected": "函数输出：资质等级计算函数+降级档位=1"},
        {"caseType": "正案例", "scenario": "准入基线=一级，当前资质=特级，验证降级档位=-1", "expected": "函数输出：资质等级计算函数+降级档位=-1"},
        {"caseType": "边界案例", "scenario": "当前资质证书为空（缺失），验证返回-999", "expected": "函数输出：资质等级计算函数+降级档位=-999"},
        {"caseType": "边界案例", "scenario": "准入基线=当前资质完全匹配，降级档位=0", "expected": "函数输出：资质等级计算函数+降级档位=0"},
    ],
}

FUNC_KEYWORD_MAP = {
    "初始化": "初始化函数", "预警等级": "预警等级计算",
    "股东公司循环": "股东公司循环函数", "持股公司循环": "持股公司循环函数",
    "股东公司数据合并": "股东公司数据合并", "持股公司数据合并": "持股公司数据合并",
    "企业信息对象": "企业信息对象取值", "个人信息对象": "个人信息对象取值",
    "关联企业信息": "关联企业信息对象取值", "资质等级": "资质等级计算函数",
}


def _find_func_cases(fname):
    """v2.1.0: 查找函数对应的用例列表，支持精确匹配和关键词模糊匹配。"""
    if fname in FUNC_CASE_MAP:
        return FUNC_CASE_MAP[fname]
    for keyword, canonical_name in FUNC_KEYWORD_MAP.items():
        if keyword in fname:
            return FUNC_CASE_MAP.get(canonical_name, [])
    return []


# ---------------------------------------------------------------------------
# v2.4.2: 通用函数用例生成器 — 当 _find_func_cases 无匹配时按 logic 类型 fallback
# 解决 generate_function_test_cases 对 FUNC_CASE_MAP 的硬依赖（"架构解耦"）
# ---------------------------------------------------------------------------

def _classify_func_type(func):
    """根据函数 logic 关键词识别类型：summary / conditional / arithmetic / fetch"""
    logic = func.get("logic", "")
    name = func.get("name", "")
    if any(k in name for k in ("汇总", "综合", "综合判断")) or any(k in logic for k in ("全部通过", "部分超限", "全部超限")):
        return "summary"
    if "IF" in logic.upper() or "THEN" in logic.upper() or "ELSE" in logic.upper():
        return "conditional"
    if any(op in logic for op in ("+", "-", "*", "/")):
        return "arithmetic"
    return "fetch"


def _generate_generic_func_cases(func):
    """基于函数逻辑类型生成最小覆盖用例集。"""
    ftype = _classify_func_type(func)
    name = func.get("name", "")
    logic = func.get("logic", "")
    cases = []

    if ftype == "arithmetic":
        cases = [
            {"caseType": "正案例", "scenario": f"{name} 正常输入（各项 > 0）计算结果正确",
             "expected": f"函数输出：{name}=各项之和（按 logic 公式计算）"},
            {"caseType": "反案例", "scenario": f"{name} 所有输入项=0，验证零值计算",
             "expected": f"函数输出：{name}=0"},
            {"caseType": "边界案例", "scenario": f"{name} 部分输入项为 null/空，验证容错计算",
             "expected": f"函数输出：{name}=非空项之和（空值按 0 处理）"},
        ]
    elif ftype == "conditional":
        cases = [
            {"caseType": "正案例", "scenario": f"{name} 满足主条件分支，验证输出",
             "expected": f"函数输出：{name}=THEN 分支结果"},
            {"caseType": "反案例", "scenario": f"{name} 不满足主条件，走 ELSE 分支",
             "expected": f"函数输出：{name}=ELSE 分支结果"},
            {"caseType": "边界案例", "scenario": f"{name} 主条件临界值（刚好等于判断阈值）",
             "expected": f"函数输出：{name}=对应分支结果（临界值处理正确）"},
        ]
    elif ftype == "summary":
        cases = [
            {"caseType": "正案例", "scenario": f"{name} 所有子维度均通过",
             "expected": f"函数输出：{name}=全部通过"},
            {"caseType": "正案例", "scenario": f"{name} 部分子维度超限、部分通过",
             "expected": f"函数输出：{name}=部分超限"},
            {"caseType": "正案例", "scenario": f"{name} 所有子维度均超限",
             "expected": f"函数输出：{name}=全部超限"},
            {"caseType": "边界案例", "scenario": f"{name} 仅 1 个子维度超限",
             "expected": f"函数输出：{name}=部分超限"},
        ]
    else:  # fetch
        cases = [
            {"caseType": "正案例", "scenario": f"{name} 传入有效输入参数",
             "expected": f"函数输出：{name}=有效返回值"},
            {"caseType": "反案例", "scenario": f"{name} 传入空输入参数",
             "expected": f"函数输出：{name}=空（安全处理）"},
        ]
    return cases


def generate_function_test_cases(functions):
    test_cases = []
    tc_func_id = 0

    for func in functions:
        fname = func["name"]
        cases = _find_func_cases(fname)
        # v2.4.2: FUNC_CASE_MAP 未匹配时按 logic 类型 fallback（架构解耦）
        if not cases:
            cases = _generate_generic_func_cases(func)
        for case in cases:
            tc_func_id += 1
            test_cases.append({
                "id": f"TC_FUNC_{tc_func_id:03d}",
                "desc": f"{fname}：{case['scenario'][:30]}",
                "group": "函数", "caseType": case["caseType"],
                "scenario": case["scenario"], "expected": case["expected"],
                "params": {
                    "S_S_BIZID": f"TC_FUNC_{tc_func_id:03d}",
                    "S_S_CUSTNO": f"TC_FUNC_CUST_{tc_func_id:03d}",
                    "S_S_ORGCODE": "010117", "S_E_APIMODEL": "30",
                },
                "targetFunction": fname,
            })

    return test_cases


# ---------------------------------------------------------------------------
# 策略架构检测 (v2.4.2)
# ---------------------------------------------------------------------------

def detect_strategy_architecture(rule_sets):
    """检测策略架构类型，区分三子策略(企业/个人/关联企业)路由 vs 扁平路由"""
    if not rule_sets:
        return "unknown"
    all_routing_empty = all(
        not rs.get("bizScenarios", "") and not rs.get("enterpriseType", "")
        for rs in rule_sets
    )
    if all_routing_empty:
        return "flat"
    names = " ".join(rs.get("name", "") for rs in rule_sets)
    codes = " ".join(rs.get("code", "") for rs in rule_sets)
    has_ent = any(kw in names for kw in ("企业",)) or "Ent" in codes
    has_per = any(kw in names for kw in ("个人",)) or "per" in codes.lower()
    has_rel = any(kw in names for kw in ("关联",)) or "relate" in codes.lower()
    if has_ent and has_per and has_rel:
        return "tripartite"
    return "flat"


# ---------------------------------------------------------------------------
# 简化版分支覆盖（非三子策略架构）(v2.4.2)
# ---------------------------------------------------------------------------

def generate_simple_flow_cases(rule_sets):
    """为非三子策略架构生成简化的分支覆盖用例：每个规则集一条覆盖"""
    test_cases = []
    tc_id = 0
    for rs in rule_sets:
        if rs.get("code", "") == "alertLevelJudgment":
            continue
        tc_id += 1
        test_cases.append({
            "id": f"TC_FLOW_{tc_id:03d}",
            "desc": f"规则集覆盖：{rs.get('name', rs['code'])}",
            "group": "决策流分支覆盖", "caseType": "正案例",
            "scenario": f"验证规则集 {rs['code']} 被正确调用",
            "expected": f"规则集{rs.get('name', rs['code'])}被调用",
            "params": {
                "S_S_BIZID": f"TC_FLOW_{tc_id:03d}",
                "S_S_CUSTNO": f"TC_FLOW_CUST_{tc_id:03d}",
                "S_S_ORGCODE": "010117", "S_E_APIMODEL": "30",
            },
            "targetRuleSet": rs["code"],
        })
    return test_cases


# ---------------------------------------------------------------------------
# 简化版综合场景（非三子策略架构）(v2.4.2)
# ---------------------------------------------------------------------------

def generate_simple_combined_cases(rules, rule_sets):
    """为非三子策略架构生成简化的综合场景用例"""
    test_cases = []
    tc_id = 0
    non_alert_rs = [rs for rs in rule_sets if rs.get("code", "") != "alertLevelJudgment"]

    # 1) 每个规则集一条综合命中用例
    for rs in non_alert_rs:
        tc_id += 1
        rs_rules = [r for r in rules if r.get("ruleSetCode", "") == rs["code"]]
        rep_rules = rs_rules[:2]
        rule_desc = "、".join(f"{r['code']}({r['hitResult']})" for r in rep_rules)
        params = {
            "S_S_BIZID": f"TC_COMB_{tc_id:03d}",
            "S_S_CUSTNO": f"TC_COMB_CUST_{tc_id:03d}",
            "S_S_ORGCODE": "010117", "S_E_APIMODEL": "30",
        }
        for r in rep_rules:
            for cond in parse_expression_conditions(r["expression"], r.get("inputParamsMapped", [])):
                params[cond["field_code"]] = construct_param_value(cond, "hit")
        test_cases.append({
            "id": f"TC_COMB_{tc_id:03d}",
            "desc": f"综合场景：{rs.get('name', rs['code'])}规则命中",
            "group": "跨子策略综合", "caseType": "正案例",
            "scenario": f"验证{rs.get('name', rs['code'])}规则集正确执行",
            "expected": f"命中：{rule_desc}" if rule_desc else f"规则集{rs['code']}执行成功",
            "params": params, "isCombined": True,
        })

    # 2) 一条全安全基线
    tc_id += 1
    rs_names = "、".join(rs.get("name", rs["code"]) for rs in non_alert_rs)
    test_cases.append({
        "id": f"TC_COMB_{tc_id:03d}",
        "desc": f"全SAFE基线：所有规则集均无命中",
        "group": "跨子策略综合", "caseType": "正案例",
        "scenario": f"所有参数安全值，验证{rs_names}均不触发",
        "expected": f"{rs_names}均无命中，最终通过",
        "params": {
            "S_S_BIZID": f"TC_COMB_{tc_id:03d}",
            "S_S_CUSTNO": f"TC_COMB_CUST_{tc_id:03d}",
            "S_S_ORGCODE": "010117", "S_E_APIMODEL": "30",
        },
        "isSafeBaseline": True,
    })

    return test_cases


# ---------------------------------------------------------------------------
# v2.4.2: 规则集联动矩阵（Phase 1-4）
# ---------------------------------------------------------------------------

def generate_ruleset_cross_matrix(rule_sets, rules, fields_dict):
    """生成 2^n 规则集命中/未命中交叉组合用例（扁平架构专用）。

    对 n 个规则集生成所有 hit/miss 组合，跳过已在单规则集覆盖和全安全基线中覆盖的组合。
    """
    from itertools import product as iter_product

    non_alert_rs = [rs for rs in rule_sets if rs.get("code", "") != "alertLevelJudgment"]
    n = len(non_alert_rs)
    if n < 2:
        return []

    # 构建 ruleSet → rule 映射（按名称推断）
    rs_to_rules = {}
    for rs in non_alert_rs:
        rs_code = rs["code"]
        rs_name = rs.get("name", "")
        matched = []
        for r in rules:
            # 按名称关键词匹配
            if "单户" in rs_name and "单户" in r.get("name", r.get("expression", "")):
                matched.append(r)
            elif "关联户" in rs_name and "关联户" in r.get("name", r.get("expression", "")):
                matched.append(r)
            elif "区域" in rs_name and "区域" in r.get("name", r.get("expression", "")):
                matched.append(r)
        if not matched:
            # fallback: 按规则编号顺序
            matched = [r for r in rules if r["code"] not in
                       sum([v2 for v2 in rs_to_rules.values()], [])]
        rs_to_rules[rs_code] = matched

    # 字段编码查找
    def _fc(name):
        """按字段名查 code，支持前缀模糊匹配。"""
        if name in fields_dict:
            return fields_dict[name].get("code")
        for k, v in fields_dict.items():
            if k.startswith(name) and len(k) > len(name):
                return v.get("code")
        return None

    # 每个维度的 hit/miss 参数模板
    dim_params = {}
    for rs in non_alert_rs:
        rs_code = rs["code"]
        rs_rules = rs_to_rules.get(rs_code, [])
        hit_p, miss_p = {}, {}

        if "单户" in rs.get("name", ""):
            fc_bal = _fc("单户合并责任余额") or "presinglemergedliability"
            fc_lim = _fc("限额_离线单户集中度") or "creditlimitoff"
            hit_p = {fc_bal: 5000000, fc_lim: 1000000}
            miss_p = {fc_bal: 100000, fc_lim: 1000000}
        elif "关联户" in rs.get("name", ""):
            fc_bal = _fc("关联户合并责任余额") or "prerelatedmergedliability"
            fc_lim = _fc("限额_离线关联户集中度") or "relatedcreditlimitoff"
            hit_p = {fc_bal: 5000000, fc_lim: 1000000}
            miss_p = {fc_bal: 100000, fc_lim: 1000000}
        elif "区域" in rs.get("name", ""):
            fc_area = _fc("区域合并已用额度") or "preareamergedusedquota"
            fc_lim = _fc("按集中度管理策略限额标准") or "conclimitstandardoff"
            fc_prov = _fc("客户区域_省") or "areaprovince"
            hit_p = {fc_area: 5000000, fc_lim: 1000000, fc_prov: "广东"}
            miss_p = {fc_area: 100000, fc_lim: 1000000, fc_prov: "广东"}

        dim_params[rs_code] = {"hit": hit_p, "miss": miss_p}

    # 生成所有 2^n 组合
    test_cases = []
    tc_id = 0
    rs_codes = [rs["code"] for rs in non_alert_rs]
    for combo in iter_product(["hit", "miss"], repeat=n):
        n_hit = sum(1 for c in combo if c == "hit")
        # 跳过已有覆盖：0 hit（全安全基线）和 1 hit（单规则集覆盖）
        if n_hit <= 1:
            continue

        tc_id += 1
        params = {
            "S_S_BIZID": f"TC_CROSS_{tc_id:03d}",
            "S_S_CUSTNO": f"TC_CROSS_CUST_{tc_id:03d}",
            "S_S_ORGCODE": "010117", "S_E_APIMODEL": "30",
        }
        hit_names, miss_names = [], []
        for i, (rs_code, state) in enumerate(zip(rs_codes, combo)):
            p = dim_params.get(rs_code, {}).get(state, {})
            params.update(p)
            rs_name = non_alert_rs[i].get("name", rs_code)
            if state == "hit":
                hit_names.append(rs_name)
            else:
                miss_names.append(rs_name)

        hit_desc = "+".join(hit_names)
        miss_desc = "+".join(miss_names) if miss_names else "无"
        combo_label = "×".join(
            f"{non_alert_rs[i].get('name', rs_codes[i])}{'命中' if c == 'hit' else '未命中'}"
            for i, c in enumerate(combo)
        )

        test_cases.append({
            "id": f"TC_CROSS_{tc_id:03d}",
            "desc": f"联动矩阵：{combo_label}",
            "group": "规则集联动矩阵", "caseType": "正案例",
            "scenario": f"规则集交叉组合：命中[{hit_desc}]，未命中[{miss_desc}]",
            "expected": f"命中：{hit_desc}；未命中：{miss_desc}",
            "params": params, "isCrossMatrix": True,
        })

    return test_cases


# ---------------------------------------------------------------------------
# v2.4.2: 区域特化模板（Phase 1-5）
# ---------------------------------------------------------------------------

def generate_region_templates(rules, rule_sets, fields_dict):
    """生成区域特化模板用例：重庆 / 四川成都 / 其他省（扁平架构专用）。

    根据区域聚合层级判断函数的分支逻辑，为每个区域生成针对性测试用例。
    """
    non_alert_rs = [rs for rs in rule_sets if rs.get("code", "") != "alertLevelJudgment"]

    def _fc(name):
        if name in fields_dict:
            return fields_dict[name].get("code")
        for k, v in fields_dict.items():
            if k.startswith(name) and len(k) > len(name):
                return v.get("code")
        return None

    fc_area = _fc("区域合并已用额度") or "preareamergedusedquota"
    fc_lim_area = _fc("按集中度管理策略限额标准") or "conclimitstandardoff"
    fc_prov = _fc("客户区域_省") or "areaprovince"
    fc_city = _fc("客户区域_市") or "areacity"
    fc_single_bal = _fc("单户合并责任余额") or "presinglemergedliability"
    fc_single_lim = _fc("限额_离线单户集中度") or "creditlimitoff"
    fc_related_bal = _fc("关联户合并责任余额") or "prerelatedmergedliability"
    fc_related_lim = _fc("限额_离线关联户集中度") or "relatedcreditlimitoff"

    # 找到 areaConcPre 对应的规则集
    area_rs = next((rs for rs in non_alert_rs if "区域" in rs.get("name", "")), None)
    area_rs_code = area_rs["code"] if area_rs else "areaConcPre"

    templates = [
        {
            "name": "重庆",
            "desc": "重庆区域模板",
            "region_params": {fc_prov: "重庆", fc_city: "重庆"},
            "target_rule": "R004",
            "rule_note": "R004条件1(省=重庆)触发",
        },
        {
            "name": "四川成都",
            "desc": "四川成都区域模板",
            "region_params": {fc_prov: "四川", fc_city: "成都"},
            "target_rule": "R004",
            "rule_note": "R004条件2(省=四川且市=成都)触发",
        },
        {
            "name": "广东",
            "desc": "其他省区域模板（非重庆、非四川成都）",
            "region_params": {fc_prov: "广东", fc_city: "广州"},
            "target_rule": "R003",
            "rule_note": "R003触发（省≠重庆，省≠四川）",
        },
    ]

    test_cases = []
    tc_id = 0

    for tmpl in templates:
        # 每条区域模板生成 2 个用例：区域命中 + 区域未命中
        for area_state in ["hit", "miss"]:
            tc_id += 1
            params = {
                "S_S_BIZID": f"TC_REGION_{tc_id:03d}",
                "S_S_CUSTNO": f"TC_REGION_CUST_{tc_id:03d}",
                "S_S_ORGCODE": "010117", "S_E_APIMODEL": "30",
            }
            params.update(tmpl["region_params"])

            # 单户和关联户都设为 hit（不影响区域维度）
            params[fc_single_bal] = 5000000
            params[fc_single_lim] = 1000000
            params[fc_related_bal] = 5000000
            params[fc_related_lim] = 1000000

            if area_state == "hit":
                params[fc_area] = 5000000
                params[fc_lim_area] = 1000000
                expected_desc = f"区域集中度超限({tmpl['rule_note']})"
                case_type = "正案例"
            else:
                params[fc_area] = 100000
                params[fc_lim_area] = 1000000
                expected_desc = "区域集中度未超限（额度低于阈值）"
                case_type = "反案例"

            test_cases.append({
                "id": f"TC_REGION_{tc_id:03d}",
                "desc": f"区域模板[{tmpl['name']}]：区域{'命中' if area_state == 'hit' else '未命中'}",
                "group": "区域特化模板", "caseType": case_type,
                "scenario": f"{tmpl['desc']}，区域集中度{'超限' if area_state == 'hit' else '安全'}",
                "expected": expected_desc,
                "params": params,
                "targetRuleSet": area_rs_code,
                "isRegionTemplate": True,
            })

    return test_cases


# ---------------------------------------------------------------------------
# 决策流分支覆盖用例
# ---------------------------------------------------------------------------

def generate_flow_coverage_cases(strategies, rule_sets):
    test_cases = []
    tc_flow_id = 0

    for biz in ["借款", "债券", "委贷", "投资", "非融"]:
        for ent in ["国企", "民营"]:
            tc_flow_id += 1
            matching_rs = []
            for rs in rule_sets:
                biz_match = biz in rs["bizScenarios"] or "全部" in rs["bizScenarios"]
                ent_match = ent in rs["enterpriseType"] or "通用" in rs["enterpriseType"] or "主策略" in rs["enterpriseType"]
                if biz_match and ent_match:
                    matching_rs.append(rs["code"])

            test_cases.append({
                "id": f"TC_FLOW_{tc_flow_id:03d}",
                "desc": f"路由覆盖：{biz}×{ent}",
                "group": "决策流分支覆盖", "caseType": "正案例",
                "scenario": f"业务类型={biz}，企业类型={ent}，验证正确的规则集被触发",
                "expected": f"路由到：{', '.join(matching_rs) if matching_rs else '主策略'}",
                "params": {
                    "S_S_BIZID": f"TC_FLOW_{tc_flow_id:03d}",
                    "S_S_CUSTNO": f"TC_FLOW_CUST_{tc_flow_id:03d}",
                    "S_S_ORGCODE": "010117", "S_E_APIMODEL": "30",
                    "C_S_BIZSCENARIO": f"{biz}类", "S_S_ENTERPRISETYPE": ent,
                },
                "targetBizType": biz, "targetEntType": ent,
                "matchedRuleSets": matching_rs,
            })

    for biz in ["贷前借款", "贷前债券"]:
        tc_flow_id += 1
        test_cases.append({
            "id": f"TC_FLOW_{tc_flow_id:03d}",
            "desc": f"路由覆盖：{biz}（贷前主策略）",
            "group": "决策流分支覆盖", "caseType": "正案例",
            "scenario": f"事件类型=贷前，业务类型={biz.replace('贷前', '')}",
            "expected": "路由到：bhjcpostMainBefore",
            "params": {"S_S_BIZID": f"TC_FLOW_{tc_flow_id:03d}", "S_S_CUSTNO": f"TC_FLOW_CUST_{tc_flow_id:03d}",
                       "S_S_ORGCODE": "010117", "S_E_APIMODEL": "30"},
            "targetBizType": biz,
        })

    special_flows = [
        {"desc": "无效企业类型(外资) DEFAULT到End", "scenario": "企业类型=外资，验证企业子策略走DEFAULT分支", "expected": "企业子策略DEFAULT分支; 个人/关联企业子策略正常", "params_extra": {"S_S_ENTERPRISETYPE": "外资", "C_S_BIZSCENARIO": "借款类"}},
        {"desc": "未知业务场景走DEFAULT路径", "scenario": "C_S_BIZSCENARIO=未知类", "expected": "三个子策略均走DEFAULT", "params_extra": {"S_S_ENTERPRISETYPE": "国企", "C_S_BIZSCENARIO": "未知类"}},
        {"desc": "投资类×国企×资管类特殊分支", "scenario": "投资+国企+资管类", "expected": "企业子策略(国企投资类资管分支)被调用", "params_extra": {"S_S_ENTERPRISETYPE": "国企", "C_S_BIZSCENARIO": "投资类", "S_S_SVTYP": "资产管理类"}},
        {"desc": "投资类×民营×资管类特殊分支", "scenario": "投资+民营+资管类", "expected": "企业子策略(民营投资类资管分支)被调用", "params_extra": {"S_S_ENTERPRISETYPE": "民营", "C_S_BIZSCENARIO": "投资类", "S_S_SVTYP": "资产管理类"}},
        {"desc": "投资类×国企×非资管类", "scenario": "投资+国企+非资管类", "expected": "企业子策略(国企投资类非资管分支)被调用", "params_extra": {"S_S_ENTERPRISETYPE": "国企", "C_S_BIZSCENARIO": "投资类", "S_S_SVTYP": "融资类"}},
        {"desc": "投资类×民营×非资管类", "scenario": "投资+民营+非资管类", "expected": "企业子策略(民营投资类非资管分支)被调用", "params_extra": {"S_S_ENTERPRISETYPE": "民营", "C_S_BIZSCENARIO": "投资类", "S_S_SVTYP": "融资类"}},
        {"desc": "借款类×国企不进入PLEP", "scenario": "个人子策略：借款×国企", "expected": "不进入PLEP规则集(仅民营)", "params_extra": {"S_S_ENTERPRISETYPE": "国企", "C_S_BIZSCENARIO": "借款类"}},
        {"desc": "个人子策略DEFAULT分支", "scenario": "非标准业务场景", "expected": "个人子策略DEFAULT分支", "params_extra": {"S_S_ENTERPRISETYPE": "国企", "C_S_BIZSCENARIO": "其他类"}},
        {"desc": "委贷类×民营: 循环不触发", "scenario": "委贷+民营，无持股数据", "expected": "循环函数不触发; 其他子策略正常", "params_extra": {"S_S_ENTERPRISETYPE": "民营", "C_S_BIZSCENARIO": "委贷类"}},
    ]

    for sf in special_flows:
        tc_flow_id += 1
        params = {"S_S_BIZID": f"TC_FLOW_{tc_flow_id:03d}", "S_S_CUSTNO": f"TC_FLOW_CUST_{tc_flow_id:03d}",
                  "S_S_ORGCODE": "010117", "S_E_APIMODEL": "30"}
        params.update(sf.get("params_extra", {}))
        test_cases.append({
            "id": f"TC_FLOW_{tc_flow_id:03d}", "desc": sf["desc"],
            "group": "决策流分支覆盖", "caseType": "正案例",
            "scenario": sf["scenario"], "expected": sf["expected"], "params": params,
        })

    return test_cases


# ---------------------------------------------------------------------------
# 策略预警等级覆盖用例
# ---------------------------------------------------------------------------

def generate_alert_level_cases(alert_rules):
    test_cases = []
    level_cases = [
        {"id": "TC_MAIN01_A", "desc": "红色预警：极高风险>0", "caseType": "正案例", "scenario": "极高风险规则命中≥1条", "expected": "命中：MAIN01(命中红色预警)[红色预警]", "hitCounts": {"极高风险": 1, "高风险": 0, "中风险": 0, "低风险": 0}},
        {"id": "TC_MAIN01_B", "desc": "红色预警：高风险>0", "caseType": "正案例", "scenario": "高风险规则命中≥1条", "expected": "命中：MAIN01(命中红色预警)[红色预警]", "hitCounts": {"极高风险": 0, "高风险": 1, "中风险": 0, "低风险": 0}},
        {"id": "TC_MAIN01_C", "desc": "红色预警：中风险≥3", "caseType": "正案例", "scenario": "中风险命中3条", "expected": "命中：MAIN01(命中红色预警)[红色预警]", "hitCounts": {"极高风险": 0, "高风险": 0, "中风险": 3, "低风险": 0}},
        {"id": "TC_MAIN01_BOUND", "desc": "红色预警边界：中风险恰好3条", "caseType": "边界案例", "scenario": "中风险=3", "expected": "命中：MAIN01(命中红色预警)[红色预警]", "hitCounts": {"极高风险": 0, "高风险": 0, "中风险": 3, "低风险": 0}},
        {"id": "TC_MAIN02_A", "desc": "黄色预警：中风险=1", "caseType": "正案例", "scenario": "高风险=0，中风险=1", "expected": "命中：MAIN02(命中黄色预警)[黄色预警]", "hitCounts": {"极高风险": 0, "高风险": 0, "中风险": 1, "低风险": 0}},
        {"id": "TC_MAIN02_B", "desc": "黄色预警：中风险=2", "caseType": "正案例", "scenario": "高风险=0，中风险=2", "expected": "命中：MAIN02(命中黄色预警)[黄色预警]", "hitCounts": {"极高风险": 0, "高风险": 0, "中风险": 2, "低风险": 0}},
        {"id": "TC_MAIN02_BOUND", "desc": "黄色预警边界：中风险=2", "caseType": "边界案例", "scenario": "中风险=2", "expected": "命中：MAIN02(命中黄色预警)[黄色预警]", "hitCounts": {"极高风险": 0, "高风险": 0, "中风险": 2, "低风险": 0}},
        {"id": "TC_MAIN03_A", "desc": "蓝色预警：仅低风险>0", "caseType": "正案例", "scenario": "仅低风险命中", "expected": "命中：MAIN03(命中蓝色预警)[蓝色预警]", "hitCounts": {"极高风险": 0, "高风险": 0, "中风险": 0, "低风险": 1}},
        {"id": "TC_MAIN04_A", "desc": "绿色预警：全部=0", "caseType": "正案例", "scenario": "所有等级=0", "expected": "命中：MAIN04(命中绿色预警)[绿色预警]", "hitCounts": {"极高风险": 0, "高风险": 0, "中风险": 0, "低风险": 0}},
        {"id": "TC_MAIN_MISS", "desc": "无任何预警", "caseType": "反案例", "scenario": "所有参数安全值", "expected": "未命中：所有预警等级规则", "hitCounts": {"极高风险": 0, "高风险": 0, "中风险": 0, "低风险": 0}},
    ]

    for case in level_cases:
        test_cases.append({
            "id": case["id"], "desc": case["desc"],
            "group": "策略预警等级覆盖", "caseType": case["caseType"],
            "scenario": case["scenario"], "expected": case["expected"],
            "params": {"S_S_BIZID": case["id"], "S_S_CUSTNO": f"CUST_{case['id']}",
                       "S_S_ORGCODE": "010117", "S_E_APIMODEL": "30"},
            "targetRuleSet": "alertLevelJudgment", "hitCounts": case["hitCounts"],
        })

    return test_cases


# ---------------------------------------------------------------------------
# 跨子策略综合场景
# ---------------------------------------------------------------------------

def generate_combined_scenarios(rules, rule_sets):
    test_cases = []
    tc_comb_id = 0

    rs_rules = {}
    for r in rules:
        rsc = r["ruleSetCode"]
        if rsc == "alertLevelJudgment":
            continue
        rs_rules.setdefault(rsc, []).append(r)

    ent_rs, per_rs, rel_rs = {}, {}, {}
    for rs in rule_sets:
        code, name = rs["code"], rs.get("name", "")
        if "个人" in name or "per" in code.lower(): per_rs[code] = rs
        elif "关联" in name or "relate" in code.lower(): rel_rs[code] = rs
        elif "企业" in name or "Ent" in code: ent_rs[code] = rs

    combos = [
        {"biz": "借款", "ent": "国企", "desc": "综合场景: 借款类国企多规则同时命中"},
        {"biz": "借款", "ent": "民营", "desc": "综合场景: 借款类民营多规则同时命中"},
        {"biz": "非融", "ent": "民营", "desc": "综合场景: 非融民营资质+中标信息"},
        {"biz": "委贷", "ent": "国企", "desc": "综合场景: 委贷类国企终本记录"},
        {"biz": "投资", "ent": "民营", "desc": "综合场景: 投资类民营破产清算"},
        {"biz": "债券", "ent": "民营", "desc": "综合场景: 债券类民营多规则命中"},
    ]

    for combo in combos:
        tc_comb_id += 1
        biz, ent = combo["biz"], combo["ent"]

        matched_ent_rs = []
        for rs_code, rs in ent_rs.items():
            biz_match = biz in rs.get("bizScenarios", "") or "全部" in rs.get("bizScenarios", "")
            ent_match = ent in rs.get("enterpriseType", "") or "通用" in rs.get("enterpriseType", "")
            if biz_match and ent_match:
                matched_ent_rs.append(rs_code)

        rep_rules = []
        for rs_code in matched_ent_rs[:3]:
            rl = rs_rules.get(rs_code, [])
            if rl: rep_rules.append(rl[0])
        for rs_code in list(per_rs.keys())[:1]:
            rl = rs_rules.get(rs_code, [])
            if rl: rep_rules.append(rl[0])
        for rs_code in list(rel_rs.keys())[:1]:
            rl = rs_rules.get(rs_code, [])
            if rl: rep_rules.append(rl[0])

        rule_desc = "、".join(f"{r['code']}({r['hitResult']})" for r in rep_rules[:4])
        params = {"S_S_BIZID": f"TC_COMB_{tc_comb_id:03d}", "S_S_CUSTNO": f"TC_COMB_CUST_{tc_comb_id:03d}",
                  "S_S_ORGCODE": "010117", "S_E_APIMODEL": "30",
                  "C_S_BIZSCENARIO": f"{biz}类", "S_S_ENTERPRISETYPE": ent}

        for r in rep_rules[:4]:
            for cond in parse_expression_conditions(r["expression"], r.get("inputParamsMapped", [])):
                params[cond["field_code"]] = construct_param_value(cond, "hit")

        test_cases.append({
            "id": f"TC_COMB_{tc_comb_id:03d}", "desc": combo["desc"],
            "group": "跨子策略综合", "caseType": "正案例",
            "scenario": f"跨子策略综合：{biz}类×{ent}，同时命中{rule_desc}",
            "expected": f"多规则同时命中：{rule_desc}，验证整体预警等级",
            "params": params, "isCombined": True,
        })

    # SAFE基线 — 填充实际安全参数值
    for biz in ["借款", "债券", "委贷", "投资", "非融"]:
        for ent in ["国企", "民营"]:
            tc_comb_id += 1
            safe_params = {"S_S_BIZID": f"TC_COMB_{tc_comb_id:03d}", "S_S_CUSTNO": f"TC_COMB_CUST_{tc_comb_id:03d}",
                           "S_S_ORGCODE": "010117", "S_E_APIMODEL": "30",
                           "C_S_BIZSCENARIO": f"{biz}类", "S_S_ENTERPRISETYPE": ent}
            safe_params.update(SAFE_DEFAULTS)

            test_cases.append({
                "id": f"TC_COMB_{tc_comb_id:03d}",
                "desc": f"全SAFE基线：{biz}×{ent} - 三个子策略均无命中",
                "group": "跨子策略综合", "caseType": "正案例",
                "scenario": f"全SAFE：{biz}类×{ent}，所有参数安全值，验证三个子策略均被调用且不触发任何规则",
                "expected": "企业子策略+个人子策略+关联企业子策略均被调用，无规则命中，最终绿色预警",
                "params": safe_params, "isSafeBaseline": True,
            })

    return test_cases


# ---------------------------------------------------------------------------
# v2.1.0: 三方数据 mock 描述（结构化对象格式）
# ---------------------------------------------------------------------------

FIELD_TO_DATASOURCE = {
    "法定代表人变更": {"code": "2003", "name": "企查查客户身份识别", "field": "ChangeList"},
    "股权变更": {"code": "2003", "name": "企查查客户身份识别", "field": "ChangeList"},
    "经营范围变更": {"code": "2003", "name": "企查查客户身份识别", "field": "ChangeList"},
    "注册资本减少": {"code": "2003", "name": "企查查客户身份识别", "field": "ChangeList"},
    "注册资本增加": {"code": "2003", "name": "企查查客户身份识别", "field": "ChangeList"},
    "注销": {"code": "2003", "name": "企查查客户身份识别", "field": "CancelInfo"},
    "失信": {"code": "740", "name": "企查查失信记录", "field": ""},
    "被执行": {"code": "741", "name": "企查查被执行人记录", "field": ""},
    "限制高消费": {"code": "742", "name": "企查查限制高消费", "field": ""},
    "终本": {"code": "761", "name": "企查查终本案件", "field": ""},
    "被保全": {"code": "887", "name": "企查查司法案件", "field": ""},
    "为被告": {"code": "887", "name": "企查查司法案件", "field": ""},
    "自诉": {"code": "887", "name": "企查查司法案件", "field": "原告"},
    "破产": {"code": "761", "name": "企查查破产案件", "field": ""},
    "欠税": {"code": "757", "name": "企查查欠税公告", "field": ""},
    "税收违法": {"code": "756", "name": "企查查税收违法", "field": ""},
    "股权冻结": {"code": "752", "name": "企查查股权冻结", "field": ""},
    "股权出质": {"code": "751", "name": "企查查股权出质", "field": ""},
    "严重违法": {"code": "748", "name": "企查查严重违法", "field": ""},
    "司法拍卖": {"code": "744", "name": "企查查司法拍卖", "field": ""},
    "行政处罚": {"code": "865", "name": "企查查行政处罚", "field": ""},
    "环保处罚": {"code": "746", "name": "企查查环保处罚", "field": ""},
    "经营异常": {"code": "739", "name": "企查查经营异常", "field": ""},
    "资质": {"code": "255", "name": "企查查资质证书", "field": ""},
    "中标": {"code": "EBNP08", "name": "中标信息查询", "field": ""},
    "实缴比例": {"code": "2003", "name": "企查查客户身份识别", "field": "PaidCapRatio"},
    "注册资本": {"code": "2003", "name": "企查查客户身份识别", "field": "RegistCapi"},
    "实缴资本": {"code": "2003", "name": "企查查客户身份识别", "field": "PaidCapi"},
    "纳税信用": {"code": "2003", "name": "企查查客户身份识别", "field": "TaxPayerLevel"},
    "成立日期": {"code": "2003", "name": "企查查客户身份识别", "field": "EstDate"},
    "股东": {"code": "2003", "name": "企查查客户身份识别", "field": "MajorShareholder"},
    "逾期借贷": {"code": "征信", "name": "征信报告", "field": ""},
    "关注类借贷": {"code": "征信", "name": "征信报告", "field": ""},
    "逾期担保": {"code": "征信", "name": "征信报告", "field": ""},
    "关注类担保": {"code": "征信", "name": "征信报告", "field": ""},
    "个人失信": {"code": "740", "name": "企查查失信记录", "field": "个人"},
    "个人被执行": {"code": "741", "name": "企查查被执行人记录", "field": "个人"},
    "个人限制高消费": {"code": "742", "name": "企查查限制高消费", "field": "个人"},
    "个人终本": {"code": "761", "name": "企查查终本案件", "field": "个人"},
    "关联注销": {"code": "2003", "name": "企查查客户身份识别", "field": "关联企业"},
    "负面舆情": {"code": "yjtxwyq", "name": "新闻舆情数据源", "field": ""},
}


def generate_third_party_mock(rule):
    """v2.1.0: 生成结构化 thirdPartyMock 对象。"""
    name = rule.get("name", "")
    expression = rule.get("expression", "")
    input_params = rule.get("inputParamsMapped", [])

    interfaces = set()
    fields = []
    hints = []

    for keyword, ds_info in FIELD_TO_DATASOURCE.items():
        if keyword in name or keyword in expression:
            interfaces.add(f"{ds_info['code']}-{ds_info['name']}")
            hint = f"mock {keyword}命中数据"
            if ds_info.get("field"):
                hint += f"({ds_info['field']})"
            hints.append(hint)

    for p in input_params:
        if p.get("code"):
            fields.append(p["code"])

    if not interfaces:
        return {"interfaces": [], "fields": fields[:10], "mockHint": "无三方数据依赖"}

    return {
        "interfaces": sorted(interfaces),
        "fields": fields[:10],
        "mockHint": "; ".join(hints[:3]),
    }


# ---------------------------------------------------------------------------
# v2.3.1: Excel 报告输出
# ---------------------------------------------------------------------------

_EXCEL_COL_NAMES = [
    '用例编号', '模块', '用例名称', '业务描述', '测试场景',
    '用例类型', '前置条件', '测试步骤', '预期结果', '实际结果',
    '测试状态', '测试时间', '测试流水号UUID', '测试数据', '三方数据', '备注'
]

_EXCEL_COL_WIDTHS = [10, 18, 40, 45, 35, 12, 15, 20, 45, 15, 10, 12, 36, 50, 40, 20]


def _format_params(params_dict):
    """将 params 字典格式化为可读文本"""
    lines = []
    for k, v in params_dict.items():
        v_str = str(v)
        if len(v_str) > 200:
            v_str = v_str[:200] + '...'
        lines.append(f'{k} = {v_str}')
    return '\n'.join(lines)


def _format_mock(mock_dict):
    """将 thirdPartyMock 格式化为可读文本"""
    if not mock_dict:
        return ''
    parts = []
    if mock_dict.get('interfaces'):
        parts.append('接口: ' + ', '.join(mock_dict['interfaces']))
    if mock_dict.get('fields'):
        parts.append('字段: ' + ', '.join(mock_dict['fields']))
    if mock_dict.get('mockHint'):
        parts.append('提示: ' + mock_dict['mockHint'])
    return '\n'.join(parts)


def export_excel(output_data, excel_path):
    """v2.3.1: 将生成结果导出为 Excel 测试报告。

    Args:
        output_data: generate_all 返回的完整输出字典
        excel_path: 输出 .xlsx 路径
    """
    if not HAS_OPENPYXL:
        print("警告：未安装 openpyxl，跳过 Excel 导出。安装命令: pip install openpyxl", file=sys.stderr)
        return None

    cases = output_data['testCases']
    strategy_name = output_data.get('strategyName', '策略测试')
    strategy_code = output_data.get('strategyCode', '')
    summary = output_data.get('summary', {})
    by_module = summary.get('byModule', {})
    by_type = summary.get('byCaseType', {})

    wb = Workbook()

    # ── Sheet 1: 汇总 ──
    ws_sum = wb.active
    ws_sum.title = '汇总'

    summary_rows = [
        ('策略名称', strategy_name),
        ('策略编码', strategy_code),
        ('生成时间', output_data.get('generatedAt', '')),
        ('生成版本', 'v2.3.1'),
        ('用例总数', summary.get('total', len(cases))),
        ('', ''),
        ('按模块分布', ''),
    ]
    for mod, cnt in by_module.items():
        summary_rows.append((mod, cnt))
    summary_rows.append(('', ''))
    summary_rows.append(('按用例类型分布', ''))
    for typ, cnt in by_type.items():
        summary_rows.append((typ, cnt))

    sum_font_bold = Font(name='微软雅黑', size=10, bold=True)
    sum_font = Font(name='微软雅黑', size=10)
    for ri, (k, v) in enumerate(summary_rows, 1):
        ws_sum.cell(row=ri, column=1, value=k).font = sum_font_bold
        ws_sum.cell(row=ri, column=2, value=v).font = sum_font
    ws_sum.column_dimensions['A'].width = 20
    ws_sum.column_dimensions['B'].width = 40

    # ── Sheet 2: 用例明细 ──
    ws = wb.create_sheet(title=strategy_name[:31])

    hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    hdr_font = Font(name='微软雅黑', size=9, color='FFFFFF', bold=True)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    data_font = Font(name='微软雅黑', size=9)
    data_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # 表头
    for ci, col_name in enumerate(_EXCEL_COL_NAMES, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = thin

    # 列宽
    for ci, w in enumerate(_EXCEL_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # 数据行
    for ri, tc in enumerate(cases, 2):
        rule_info = f"{tc.get('targetRuleSet', '')}/{tc.get('targetRule', '')}"
        row = [
            tc.get('id', ''),
            tc.get('group', ''),
            tc.get('desc', ''),
            tc.get('scenario', ''),
            rule_info,
            tc.get('caseType', ''),
            '',
            '提交策略测试，验证执行结果',
            tc.get('expected', ''),
            '待执行', '待执行', '待执行', '待执行',
            _format_params(tc.get('params', {})),
            _format_mock(tc.get('thirdPartyMock', {})),
            '',
        ]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(_EXCEL_COL_NAMES))}{len(cases) + 1}'

    wb.save(excel_path)
    print(f"Excel 报告已导出: {excel_path}", file=sys.stderr)
    return excel_path


# ---------------------------------------------------------------------------
# 反馈注入 (v2.4.2: --feedback 支持，闭合 Agent Loop)
# ---------------------------------------------------------------------------

def load_feedback(feedback_path):
    """加载上轮反馈 JSON，返回 actions 列表"""
    if not feedback_path or not Path(feedback_path).exists():
        return []
    try:
        data = json.loads(Path(feedback_path).read_text(encoding="utf-8"))
        return data.get("actions", [])
    except (json.JSONDecodeError, KeyError) as e:
        print(f"警告：无法加载反馈文件: {e}", file=sys.stderr)
        return []


def apply_feedback(all_cases, actions):
    """
    将反馈动作应用到用例列表。匹配优先级：targetRule > scenario 关键词 > desc 关键词。

    支持的动作类型：
      fixParams      — 修正指定字段的入参值
      adjustExpected — 替换预期结果
      removeCases    — 移除无效用例

    返回应用的反馈数量。
    """
    if not actions:
        return 0

    # 构建索引
    by_rule = {}
    by_scenario_kw = {}
    for tc in all_cases:
        rule = tc.get("targetRule", "")
        if rule:
            by_rule.setdefault(rule, []).append(tc)
        scenario = tc.get("scenario", "") + " " + tc.get("desc", "")
        for word in re.findall(r'[A-Z][A-Z0-9_]{2,}', scenario):
            by_scenario_kw.setdefault(word, []).append(tc)

    applied = 0
    for action in actions:
        atype = action.get("type", "")
        rule = action.get("rule", "")
        target = action.get("target", "")
        field = action.get("field", "")
        reason = action.get("reason", "")

        # 尝试按规则/函数代码匹配
        matches = by_rule.get(rule, [])

        # 回退：从 reason 中提取规则代码
        if not matches and not rule:
            rule_match = re.search(
                r'(ARGP\d+|EBGS\d+|EBGP\d+|EBNP\d+|EBNS\d+|ENDX\d+|ARMP\d+|'
                r'ELLP\d+|ELLS\d+|MAIN\d+|S\d{6})',
                reason
            )
            if rule_match:
                matches = by_rule.get(rule_match.group(1), [])

        # 回退：按场景关键词匹配
        if not matches:
            for word in re.findall(r'[A-Z][A-Z0-9_]{2,}', reason):
                if word in by_scenario_kw:
                    matches = by_scenario_kw[word]
                    break

        if not matches:
            continue

        tc = matches[0]
        if atype == "fixParams" and field and field != "_auto_detect":
            new_val = action.get("newValue")
            if new_val is not None:
                tc.setdefault("params", {})[field] = new_val
                applied += 1
        elif atype == "fixParams" and field == "_auto_detect":
            # 无法自动确定具体字段，标记诊断信息供下轮参考
            tc.setdefault("_feedback_pending", []).append({
                'reason': action.get('reason', ''),
                'action': action.get('action', ''),
                'rule': action.get('rule', ''),
            })
            applied += 1
        elif atype == "adjustExpected":
            new_exp = action.get("newValue")
            if new_exp:
                tc["expected"] = new_exp
                applied += 1
        elif atype == "removeCases":
            if tc in all_cases:
                all_cases.remove(tc)
                applied += 1

    return applied


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------

def generate_all(parsed_json_path, output_path, feedback_path=None):
    """生成全部测试用例。feedback_path 可选，用于注入上轮 Agent Loop 反馈。"""
    path = Path(parsed_json_path)
    if not path.exists():
        print(f"错误：输入文件不存在: {parsed_json_path}", file=sys.stderr)
        sys.exit(1)

    try:
        raw_text = path.read_text(encoding="utf-8")
    except (UnicodeDecodeError, PermissionError) as e:
        print(f"错误：无法读取文件: {e}", file=sys.stderr)
        sys.exit(1)

    try:
        data = json.loads(raw_text)
    except json.JSONDecodeError as e:
        print(f"错误：JSON 解析失败: {e}", file=sys.stderr)
        sys.exit(1)

    required_keys = ["rules", "fields", "functions", "strategies", "ruleSets"]
    missing = [k for k in required_keys if k not in data]
    if missing:
        print(f"错误：parsed JSON 缺少必要字段: {', '.join(missing)}", file=sys.stderr)
        sys.exit(1)

    rules = data["rules"]
    functions = data["functions"]
    strategies = data["strategies"]
    rule_sets = data["ruleSets"]

    strategy_code = data.get("strategies", [{}])[0].get("code", "unknown")
    strategy_name = data.get("strategies", [{}])[0].get("name", "未知策略")

    alert_rules = [r for r in rules if r["ruleSetCode"] == "alertLevelJudgment"]

    rule_cases = generate_rule_test_cases(rules, data["fields"])

    # v2.3.1 BUG-3 fix: 规则级用例注入路由参数（C_S_BIZSCENARIO / S_S_ENTERPRISETYPE）
    # 天策策略的子策略分支依赖这两个参数路由到正确的规则集，缺失则用例无法命中目标规则
    # v2.4.2: 仅对 tripartite 架构注入；扁平架构的规则集没有 bizScenarios×enterpriseType 路由维度
    arch = detect_strategy_architecture(rule_sets)
    if arch == "tripartite":
        rs_routing = {}
        for rs in rule_sets:
            rsc = rs["code"]
            biz_list = rs.get("bizScenarios", "")
            ent_list = rs.get("enterpriseType", "")
            biz_val = biz_list.split(",")[0].strip() if biz_list and biz_list != "全部" else ""
            if ent_list:
                ent_parts = [e.strip() for e in ent_list.split(",")]
                ent_val = ent_parts[0]
                for preferred in ("国企", "民营"):
                    if preferred in ent_parts:
                        ent_val = preferred
                        break
            else:
                ent_val = ""
            if ent_val in ("通用", "主策略"):
                ent_val = "国企"
            rs_routing[rsc] = (biz_val, ent_val)

        injected_count = 0
        for tc in rule_cases:
            rs_code = tc.get("targetRuleSet", "")
            if rs_code in rs_routing:
                biz_val, ent_val = rs_routing[rs_code]
                params = tc.get("params", {})
                if biz_val and "C_S_BIZSCENARIO" not in params:
                    params["C_S_BIZSCENARIO"] = f"{biz_val}类"
                if ent_val and "S_S_ENTERPRISETYPE" not in params:
                    params["S_S_ENTERPRISETYPE"] = ent_val
                tc["params"] = params
                injected_count += 1
        if injected_count:
            print(f"已为 {injected_count} 条规则级用例注入路由参数", file=sys.stderr)
    else:
        print(f"扁平架构({arch})：跳过 C_S_BIZSCENARIO/S_S_ENTERPRISETYPE 路由参数注入", file=sys.stderr)

    func_cases = generate_function_test_cases(functions)

    # v2.4.2: 策略架构检测 — 根据架构类型选择生成逻辑
    arch = detect_strategy_architecture(rule_sets)
    if arch == "tripartite":
        flow_cases = generate_flow_coverage_cases(strategies, rule_sets)
        alert_cases = generate_alert_level_cases(alert_rules)
        combined_cases = generate_combined_scenarios(rules, rule_sets)
        cross_cases = []
        region_cases = []
    else:
        print(f"检测到扁平架构({arch})，使用简化生成器（跳过保后检查专用模板）", file=sys.stderr)
        flow_cases = generate_simple_flow_cases(rule_sets)
        alert_cases = []
        combined_cases = generate_simple_combined_cases(rules, rule_sets)
        # v2.4.2 Phase 1-4: 规则集联动矩阵
        cross_cases = generate_ruleset_cross_matrix(rule_sets, rules, data["fields"])
        if cross_cases:
            print(f"已生成 {len(cross_cases)} 条规则集联动矩阵用例 (v2.4.2)", file=sys.stderr)
        # v2.4.2 Phase 1-5: 区域特化模板
        region_cases = generate_region_templates(rules, rule_sets, data["fields"])
        if region_cases:
            print(f"已生成 {len(region_cases)} 条区域特化模板用例 (v2.4.2)", file=sys.stderr)

    all_cases = rule_cases + func_cases + flow_cases + alert_cases + combined_cases + cross_cases + region_cases

    # v2.4.2 Phase 1-6: 动态字段自检 — 确保所有用例包含 API 必传字段
    api_required = ["S_S_IDNO", "C_F_RANDOMNUM"]
    strategy_meta = data.get("strategies", [{}])[0]
    extra_required = strategy_meta.get("apiRequiredFields", [])
    if extra_required:
        api_required = list(set(api_required + extra_required))
    injected_count = 0
    for tc in all_cases:
        params = tc.get("params", {})
        for field in api_required:
            if field not in params:
                if field == "S_S_IDNO":
                    params[field] = "500102199001011234"
                elif field == "C_F_RANDOMNUM":
                    params[field] = "0.5"
                else:
                    params[field] = ""
                injected_count += 1
    if injected_count:
        print(f"已注入 {injected_count} 个 API 必传字段 (v2.4.2)", file=sys.stderr)

    # v2.4.2: 注入上轮 Agent Loop 反馈（fixParams / adjustExpected / removeCases）
    if feedback_path:
        fb_actions = load_feedback(feedback_path)
        if fb_actions:
            fb_applied = apply_feedback(all_cases, fb_actions)
            print(f"已应用 {fb_applied}/{len(fb_actions)} 条反馈动作 (v2.4.2)", file=sys.stderr)

    # v2.3.1: 边界用例归入正/反案例（按预期结果判定：命中→正案例，未命中→反案例）
    boundary_count = 0
    for tc in all_cases:
        if tc.get("caseType") == "边界案例":
            expected = tc.get("expected", "")
            if "未命中" in expected or "不命中" in expected or "不触发" in expected:
                tc["caseType"] = "反案例"
            else:
                tc["caseType"] = "正案例"
            boundary_count += 1
    if boundary_count:
        print(f"已将 {boundary_count} 条边界用例归入正/反案例", file=sys.stderr)

    for i, tc in enumerate(all_cases, start=1):
        tc["id"] = f"TC_{i:03d}"

    for tc in all_cases:
        if tc.get("targetRule"):
            for r in rules:
                if r["code"] == tc["targetRule"]:
                    tc["thirdPartyMock"] = generate_third_party_mock(r)
                    break

    by_module = {}
    by_type = {}
    for tc in all_cases:
        by_module[tc["group"]] = by_module.get(tc["group"], 0) + 1
        by_type[tc["caseType"]] = by_type.get(tc["caseType"], 0) + 1

    output = {
        "strategyCode": strategy_code, "strategyName": strategy_name,
        "generatedAt": datetime.now().isoformat(),
        "summary": {"total": len(all_cases), "byModule": by_module, "byCaseType": by_type},
        "testCases": all_cases,
    }

    Path(output_path).write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"生成完成！共 {len(all_cases)} 条测试用例 (v2.4.2)", file=sys.stderr)
    for mod, cnt in by_module.items():
        print(f"  {mod}: {cnt}", file=sys.stderr)
    print(f"  ---", file=sys.stderr)
    for typ, cnt in by_type.items():
        print(f"  {typ}: {cnt}", file=sys.stderr)
    print(f"输出到: {output_path}", file=sys.stderr)

    # v2.3.1: 同时导出 Excel 报告
    excel_path = str(Path(output_path).with_suffix('.xlsx'))
    export_excel(output, excel_path)

    return output


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="天策策略测试用例生成引擎 v2.4.2",
        epilog="示例: python3 generate_testcases.py parsed.json output.json --feedback feedback.json",
    )
    parser.add_argument("parsed_json", help="策略解析 JSON 文件路径 (parsed_strategy.json)")
    parser.add_argument("output_json", help="输出测试用例 JSON 文件路径")
    parser.add_argument("--feedback", "-f", default=None,
                        help="上轮 Agent Loop 反馈 JSON (feedback.json)，用于自动修正用例")
    parser.add_argument("--excel-output", default=None,
                        help="Excel 报告输出路径（默认与 output_json 同目录同名 .xlsx）")
    args = parser.parse_args()

    result = generate_all(args.parsed_json, args.output_json, feedback_path=args.feedback)

    # 如果指定了 --excel-output，覆盖默认的 Excel 路径
    if args.excel_output and result:
        from shutil import copy2
        default_xlsx = str(Path(args.output_json).with_suffix('.xlsx'))
        if Path(default_xlsx).exists():
            if Path(default_xlsx).resolve() != Path(args.excel_output).resolve():
                copy2(default_xlsx, args.excel_output)
                print(f"Excel 报告已复制到: {args.excel_output}", file=sys.stderr)
            else:
                print(f"Excel 报告已生成: {args.excel_output}", file=sys.stderr)
