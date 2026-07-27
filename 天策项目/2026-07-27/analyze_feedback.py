#!/usr/bin/env python3
"""
tiance-agent-loop / Step 4: 反馈分析 v1.1.0
解析 tiance-report-checker 的检查结果，分类问题根因，生成可操作的 feedback.json。

v1.1.0: JSON 损坏保护、空 issues 早退出、'提示'级别兜底处理。
v1.0.0: 初始版本。

输入:
  --check-result:  checker --json 输出的 JSON 摘要
  --checked-report: checker 标注后的 Excel（含质量检查 Sheet，可选）
  --test-report:   原始测试报告 Excel（可选，用于交叉验证）

输出:
  feedback.json — 传递给下一轮 generator 的结构化反馈
"""

import argparse
import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    openpyxl = None

# 规则编号正则
RULE_RE = re.compile(
    r'(ARGP\d+|EBGS\d+|EBGP\d+|EBNP\d+|EBNS\d+|ENDX\d+|ARMP\d+|'
    r'ELLP\d+|ELLS\d+|MAIN\d+|RED_ALERT_\w+|RULE_\w+)'
)
FUNC_RE = re.compile(r'(S\d{6}|FUNC_\d+)')


# ============================================================
# 问题分类器
# ============================================================

def classify_issues(issues):
    """
    将 checker 的 issues 按根因分类，生成动作列表。

    Returns:
        tuple: (actions list, stats dict)
    """
    actions = []
    stats = {
        'fixParams': 0, 'adjustExpected': 0, 'manualReview': 0,
        'keepAsPair': 0, 'reSubmit': 0, 'investigate': 0,
    }

    for issue in issues:
        tc_id = issue.get('tc_id', '')
        base_m = re.match(r'(TC_\d+)', tc_id)
        base_id = base_m.group(1) if base_m else tc_id
        category = issue.get('category', '')
        level = issue.get('level', '')
        desc = issue.get('issue', '')

        # 1. 判定一致性 — 目标规则未命中
        if category == '判定一致性' and '未命中' in desc:
            rule_m = RULE_RE.search(desc)
            rule_id = rule_m.group(1) if rule_m else 'unknown'
            actions.append({
                'type': 'fixParams', 'target': base_id, 'rule': rule_id,
                'field': '_auto_detect',
                'reason': f'{rule_id} 预期命中但实际未命中: {desc[:80]}',
                'action': f'检查 {rule_id} 的 mock 数据是否满足触发条件，调整入参使其达到阈值',
            })
            stats['fixParams'] += 1

        # 2. 实际结果质量 — 函数输出不匹配
        elif category == '实际结果质量' and '函数' in desc and '不匹配' in desc:
            func_m = FUNC_RE.search(desc)
            func_id = func_m.group(1) if func_m else 'unknown'
            actions.append({
                'type': 'adjustExpected', 'target': base_id, 'rule': func_id,
                'field': 'expected',
                'reason': f'函数 {func_id} 输出值与预期不一致: {desc[:80]}',
                'action': f'从执行日志读取 {func_id} 的实际返回值，修正用例预期',
            })
            stats['adjustExpected'] += 1

        # 3. 判定一致性 — 其他判定问题（严重）
        elif category == '判定一致性' and level == '严重':
            actions.append({
                'type': 'manualReview', 'target': base_id,
                'reason': f'严重判定一致性问题: {desc[:100]}',
                'action': '人工确认是验证标准问题还是策略缺陷',
            })
            stats['manualReview'] += 1

        # 4. 语义重复检测
        elif category == '语义重复':
            actions.append({
                'type': 'keepAsPair', 'target': base_id,
                'reason': f'跨子策略对应规则，属设计意图: {desc[:80]}',
                'action': '保留，不消除',
            })
            stats['keepAsPair'] += 1

        # 5. 版本不一致
        elif category == '版本不一致' or 'policyVersion' in desc:
            actions.append({
                'type': 'reSubmit', 'target': base_id,
                'reason': f'策略版本不一致: {desc[:80]}',
                'action': '使用最新 policyVersion 重新提交',
            })
            stats['reSubmit'] += 1

        # 6. 策略报异常
        elif '异常' in desc and ('报异常' in desc or '中断' in desc):
            actions.append({
                'type': 'fixParams', 'target': base_id,
                'field': '_auto_detect',
                'reason': f'策略执行异常: {desc[:80]}',
                'action': '检查入参格式（JSON 数组/对象格式是否正确）',
            })
            stats['fixParams'] += 1

        # 5. 实际结果质量 — 其他（错误关键词、风险等级不一致等）
        elif category == '实际结果质量':
            actions.append({
                'type': 'adjustExpected' if '不一致' in desc else 'manualReview',
                'target': base_id,
                'reason': f'实际结果质量问题: {desc[:80]}',
                'action': '核对实际结果与预期，修正用例预期值或检查策略逻辑',
            })
            stats['adjustExpected' if '不一致' in desc else 'manualReview'] += 1

        # 6. 数据完整性问题
        elif category == '数据完整性':
            actions.append({
                'type': 'investigate', 'target': base_id,
                'reason': f'数据缺失: {desc[:80]}',
                'action': '检查报告生成流程，确认数据是否正确写入',
            })
            stats['investigate'] += 1

        # 7. 系统回查比对
        elif category == '系统回查比对':
            if '会话' in desc and '过期' in desc:
                actions.append({
                    'type': 'reSubmit', 'target': base_id,
                    'reason': f'会话过期导致系统回查失败: {desc[:80]}',
                    'action': '更新 cookie 后重新提交测试',
                })
                stats['reSubmit'] += 1
            elif '不可达' in desc or '平台不可达' in desc:
                actions.append({
                    'type': 'investigate', 'target': base_id,
                    'reason': f'平台不可达: {desc[:80]}',
                    'action': '检查平台连通性后重试',
                })
                stats['investigate'] += 1
            elif '不一致' in desc:
                actions.append({
                    'type': 'manualReview', 'target': base_id,
                    'reason': f'系统回查不一致: {desc[:100]}',
                    'action': '人工比对系统记录与报告数据',
                })
                stats['manualReview'] += 1
            elif '执行失败' in desc and '通过' in desc:
                actions.append({
                    'type': 'manualReview', 'target': base_id,
                    'reason': f'系统执行失败但报告标记通过: {desc[:100]}',
                    'action': '人工确认是验证标准问题还是报告记录错误',
                })
                stats['manualReview'] += 1
            else:
                actions.append({
                    'type': 'investigate', 'target': base_id,
                    'reason': f'系统回查问题: {desc[:80]}',
                    'action': '检查系统记录与报告差异',
                })
                stats['investigate'] += 1

        # 8. 其他（严重/警告/提示分级处理）
        else:
            if level == '严重':
                actions.append({
                    'type': 'manualReview', 'target': base_id,
                    'reason': f'严重问题({category}): {desc[:100]}',
                    'action': '需人工判断',
                })
                stats['manualReview'] += 1
            elif level == '警告':
                actions.append({
                    'type': 'investigate', 'target': base_id,
                    'reason': f'警告({category}): {desc[:100]}',
                    'action': '建议人工检查',
                })
                stats['investigate'] += 1
            elif level == '提示':
                # 提示级别问题保留记录但不生成动作，避免噪音
                actions.append({
                    'type': 'investigate', 'target': base_id,
                    'reason': f'提示({category}): {desc[:100]}',
                    'action': '低优先级，建议有空时检查',
                })
                stats['investigate'] += 1

    return actions, stats


# ============================================================
# 聚合与去重
# ============================================================

def aggregate_feedback(actions, stats, check_result):
    by_target = defaultdict(list)
    for action in actions:
        by_target[action['target']].append(action)

    deduped = []
    for target, group in by_target.items():
        seen = set()
        for a in group:
            key = (a['type'], a.get('rule', ''), a.get('field', ''))
            if key not in seen:
                seen.add(key)
                deduped.append(a)

    priority = {'manualReview': 0, 'fixParams': 1, 'adjustExpected': 2,
                'reSubmit': 3, 'investigate': 4, 'keepAsPair': 5}
    deduped.sort(key=lambda x: (priority.get(x['type'], 9), x['target']))

    current_issues = check_result.get('total_issues', 0)
    auto_fixable = stats['fixParams'] + stats['adjustExpected'] + stats['reSubmit']
    keep = stats['keepAsPair']
    estimated_next = max(0, current_issues - auto_fixable - keep)

    return {
        'actions': deduped,
        'summary': {
            'totalActions': len(deduped),
            'autoFixable': auto_fixable,
            'manualReview': stats['manualReview'],
            'keepAsPair': keep,
            'investigate': stats['investigate'],
            'estimatedNextIssues': estimated_next,
        },
        'previousIssues': current_issues,
    }


# ============================================================
# 从 checked-report 读取 issues（当 --json 不含 issues 时使用）
# ============================================================

def read_issues_from_checked_excel(excel_path):
    """从质量检查 Sheet 读取 issues 列表（使用 iter_rows 快速读取）"""
    if openpyxl is None:
        print('警告: openpyxl 未安装，跳过 Excel 读取', file=sys.stderr)
        return []

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    issues = []
    try:
        if '质量检查' not in wb.sheetnames:
            return []
        ws = wb['质量检查']

        # 用 iter_rows 批量读取，比 cell-by-cell 快几十倍
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(v or '').strip() for v in next(rows_iter)]

        # 找到问题明细表的起始位置（跳过统计摘要部分）
        detail_start = False
        for row_vals in rows_iter:
            vals = list(row_vals)
            row_strs = [str(v or '').strip() for v in vals]

            # 检测表头行
            if not detail_start:
                if '行号' in row_strs or '用例编号' in row_strs:
                    headers = row_strs
                    detail_start = True
                continue

            # 读取数据行
            row = dict(zip(headers, row_strs))
            level = row.get('严重级别', row.get('严重程度', ''))
            if not level:
                continue

            issues.append({
                'row': int(row.get('行号', 0) or 0),
                'tc_id': row.get('用例编号', ''),
                'level': level,
                'category': row.get('检查维度', row.get('问题类别', '')),
                'issue': row.get('问题描述', ''),
            })
    finally:
        wb.close()

    return issues


# ============================================================
# CLI
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description='Agent Loop Step 4: 反馈分析 — 解析检查结果，生成优化反馈',
    )
    parser.add_argument('--check-result', required=True, help='checker --json 输出文件')
    parser.add_argument('--checked-report', help='checker 标注后的 Excel（可选）')
    parser.add_argument('--test-report', help='原始测试报告 Excel（可选，用于交叉验证）')
    parser.add_argument('--output', '-o', required=True, help='输出 feedback.json 路径')
    parser.add_argument('--iteration', type=int, default=1, help='当前迭代轮次')

    args = parser.parse_args()

    # 读取检查结果摘要（含 JSON 损坏保护）
    try:
        with open(args.check_result, 'r', encoding='utf-8') as f:
            check_result = json.load(f)
    except json.JSONDecodeError as e:
        print(f'错误: check-result JSON 损坏: {e}', file=sys.stderr)
        print(f'文件: {args.check_result}', file=sys.stderr)
        sys.exit(1)
    except FileNotFoundError:
        print(f'错误: check-result 文件不存在: {args.check_result}', file=sys.stderr)
        sys.exit(1)

    if not isinstance(check_result, dict):
        print(f'错误: check-result 格式异常，期望 dict 实际 {type(check_result).__name__}',
              file=sys.stderr)
        sys.exit(1)

    # 获取 issues 列表
    issues = check_result.get('issues', [])
    if not issues and args.checked_report and os.path.exists(args.checked_report):
        print(f'从 {args.checked_report} 读取 issues...', file=sys.stderr)
        issues = read_issues_from_checked_excel(args.checked_report)

    # 空 issues 早退出：生成最小 feedback 避免下游空跑
    if not issues:
        print('无 issues，生成空反馈...', file=sys.stderr)
        empty_feedback = {
            'iteration': args.iteration,
            'generatedAt': datetime.now().isoformat(),
            'policyCode': check_result.get('policyCode', ''),
            'previousIssues': check_result.get('total_issues', 0),
            'actions': [],
            'summary': {
                'totalActions': 0, 'autoFixable': 0, 'manualReview': 0,
                'keepAsPair': 0, 'investigate': 0, 'estimatedNextIssues': 0,
            },
        }
        os.makedirs(os.path.dirname(args.output) or '.', exist_ok=True)
        with open(args.output, 'w', encoding='utf-8') as f:
            json.dump(empty_feedback, f, ensure_ascii=False, indent=2)
        print(f'空反馈已保存: {args.output}', file=sys.stderr)
        return

    # 分类
    print(f'分析问题: {len(issues)} 条...', file=sys.stderr)
    actions, stats = classify_issues(issues)

    # 聚合
    feedback = aggregate_feedback(actions, stats, check_result)
    feedback['iteration'] = args.iteration
    feedback['generatedAt'] = datetime.now().isoformat()
    feedback['policyCode'] = check_result.get('policyCode', '')

    # 写入
    os.makedirs(os.path.dirname(args.output) or '.', exist_ok=True)
    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(feedback, f, ensure_ascii=False, indent=2)

    # 输出摘要
    s = feedback['summary']
    print(f'\n{"=" * 50}')
    print(f'  反馈分析完成 — tiance-agent-loop Step 4')
    print(f'{"=" * 50}')
    print(f'  当前问题总数:     {feedback["previousIssues"]}')
    print(f'  生成动作总数:     {s["totalActions"]}')
    print(f'  ├─ 可自动修复:    {s["autoFixable"]}')
    print(f'  ├─ 需人工确认:    {s["manualReview"]}')
    print(f'  ├─ 保留(设计意图): {s["keepAsPair"]}')
    print(f'  └─ 待调查:        {s["investigate"]}')
    print(f'  预估下轮问题数:   {s["estimatedNextIssues"]}')
    print(f'\n  反馈文件已保存:   {args.output}')


if __name__ == '__main__':
    main()
